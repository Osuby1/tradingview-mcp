#!/usr/bin/env python3
"""Curated, human-readable trading report (v3).

Replaces the raw-dump style of compile_universe_report_v2.py with a report
you can actually read top to bottom: a bottom line, a market dashboard,
sized recommendations with the reasoning spelled out, and — the part that
was missing — a WHY NOT sheet explaining every rejected name and what
would have to change for it to qualify.

Usage:  python scripts/compile_curated_report.py [YYYY-MM-DD]
Writes: reports/curated_report_<date>.xlsx
"""
import os
import sys
import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATE = sys.argv[1] if len(sys.argv) > 1 else datetime.date.today().isoformat()
WRITTEN_AT = datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M %Z")

# ---------------------------------------------------------------- palette
NAVY = "1F3864"
BLUE = "2E5C8A"
GREY = "F2F2F2"
GREEN = "C6EFCE"
RED = "FFC7CE"
AMBER = "FFF2CC"
WHITE = "FFFFFF"

H1 = Font(bold=True, size=16, color=WHITE)
H2 = Font(bold=True, size=11, color=WHITE)
BOLD = Font(bold=True, size=10)
BODY = Font(size=10)
SMALL = Font(size=9, italic=True, color="595959")
TOP = Alignment(vertical="top", wrap_text=True)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()


def sheet(title, widths, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    return ws


def banner(ws, text, ncols, sub=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font, c.fill, c.alignment = H1, PatternFill("solid", fgColor=NAVY), Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        s = ws.cell(row=2, column=1, value=sub)
        s.font, s.alignment = SMALL, Alignment(vertical="center", indent=1)
        ws.row_dimensions[2].height = 16


def header(ws, row, labels):
    for i, lab in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font, c.fill, c.alignment, c.border = H2, PatternFill("solid", fgColor=BLUE), CTR, BOX
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def row(ws, r, values, fill=None, bold_first=False):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = BOLD if (bold_first and i == 1) else BODY
        c.alignment, c.border = TOP, BOX
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)


# ============================================================ 1. START HERE
ws = sheet("START HERE", [104], first=True)
banner(ws, f"TRADING REPORT — {DATE}", 1, f"Written {WRITTEN_AT}   |   prices from the live TradingView feed")
r = 4
for text, font in [
    ("BOTTOM LINE", BOLD),
    ("One trade clears every gate today: MU. Everything else is blocked, and the Why Not sheet gives the specific "
     "number behind each rejection rather than a verdict you have to take on faith.", BODY),
    ("", BODY),
    ("The tape is not the story today — the Fed is. Polymarket's July hike tail jumped to ~19% from ~5% a week ago, "
     "almost all of it in the last 24 hours, because crude is up ~20% month-to-date and $94 Brent reads as an "
     "inflation shock six days before the FOMC. Recession odds did NOT move. The market is pricing a supply shock, "
     "not a demand shock — which favours energy, gold and the shock-beneficiaries over duration and defensives.", BODY),
    ("", BODY),
    ("The uncomfortable part: the tanker thesis is not confirming. Brent +2.8% to $94, VLCC day rates near "
     "$370k, Hormuz transits down ~90% year-over-year — and both NAT and STNG are RED and sitting at session lows, "
     "below their triggers. Either the equities are discounting the ceasefire proposal or the move is already "
     "priced. Neither has triggered. Not forcing it is the trade.", BODY),
    ("", BODY),
    ("HOW TO READ THIS", BOLD),
    ("  Market Dashboard — the tape, the Fed, oil and rotation in plain English. Read this first.", BODY),
    ("  Recommendations — sized trades only. Entry, stop, share count, dollar risk, target, and WHY.", BODY),
    ("  Why Not — every name that did NOT make it, the exact number that killed it, and what would revive it.", BODY),
    ("  Watch List — names one condition away, with the specific trigger to watch.", BODY),
    ("  Sell Mode — indicator says downtrend. Rule 10: do not buy these, regardless of how good the story is.", BODY),
    ("  Method & Data Quality — how signals were derived, and every known gap. Read before trusting a number.", BODY),
    ("", BODY),
    ("THE GATES, IN ORDER", BOLD),
    ("  1. Chandelier mode — SELL mode is an automatic veto (rule 10), no exceptions.", BODY),
    ("  2. Price vs ZLSMA — below the line means the trend has not turned yet.", BODY),
    ("  3. Magical OB/OS — above +100 is too hot, pullback-only. Below -100 is washed out.", BODY),
    ("  4. Signal freshness — the buy must be <=5 sessions old, dated from the indicator's own plot series.", BODY),
    ("  5. Earnings gate — no new position into a print inside the trade window.", BODY),
    ("  6. Risk/reward — target must be >=2x risk measured to the NEAREST overhead resistance, not to a hope.", BODY),
    ("  7. Sizing — $5,000 risk per trade, $100,000 maximum order, structure stop with a 1x ATR floor.", BODY),
    ("", BODY),
    ("A name has to pass ALL SEVEN. Most failures below are single-gate failures — the setup was otherwise fine, "
     "which is exactly why the reason matters more than the verdict.", BODY),
]:
    c = ws.cell(row=r, column=1, value=text)
    c.font, c.alignment = font, TOP
    r += 1

# ====================================================== 2. MARKET DASHBOARD
ws = sheet("Market Dashboard", [26, 20, 78])
banner(ws, "MARKET DASHBOARD", 3, "Live feed unless marked. Percentages vs prior close.")
header(ws, 3, ["What", "Level", "What it means"])
DASH = [
    ("THE FED — headline", "Hike tail ~19%", "Polymarket's July 28-29 hike odds jumped ~10.6 points in 24 hours and ~14.5 in a week, to roughly 19% including the +50bp sliver. A week ago this was a ~5% tail. Crude is repricing the Fed in real time. This is the single most important number on the page.", AMBER),
    ("2026 rate cuts", "0 cuts: 85%", "Easing is essentially priced out of 2026 entirely. The 1-cut bucket is bleeding.", None),
    ("Recession odds 2026", "10%", "FLAT while the hike tail exploded. The market reads this as an inflation shock, not a demand shock — long the shock beats short the cycle.", None),
    ("", "", "", None),
    ("Brent", "$94.05", "+2.8%, fourth straight up day, ~+20% month-to-date. This is the engine under everything else today.", AMBER),
    ("WTI", "$87.05", "+2.8%. Polymarket now has $90 WTI at ~69% (was ~48% yesterday) and $100 at ~17%. The $80 and $85 strikes both resolved YES in the last 24 hours.", None),
    ("Strait of Hormuz", "Transits -90% y/y", "Non-Iran-linked transits collapsed to 25 from 108 week-on-week. At least 9 ships attacked since July 6. VLCC TD3C assessed near $370k/day.", None),
    ("Ceasefire odds", "20% by 7/31", "The market is FADING the Qatar/Pakistan 10-day proposal — Aug 14 odds fell 3 points and Aug 31 fell 4 points over the last day. A ceasefire is not the base case before August, and confidence went DOWN overnight, not up.", None),
    ("", "", "", None),
    ("SPY", "748.01", "Flat. Futures were soft pre-open on chip weakness, into the heaviest tech earnings night of the quarter.", None),
    ("QQQ", "706.57", "Modestly lower. Caution ahead of Tesla and Alphabet tonight.", None),
    ("Leadership", "MEMORY semis", "MU +18% and SanDisk +21% over three sessions, extending again today. Read this as a memory-cycle trade, NOT a broad AI-semis rally — index futures were pressured BY chip weakness while memory led. That is rotation within semis.", GREEN),
    ("Gold", "$4,150", "+1.7%. GDX +4.1%, miners levering the move ~2.5x. Watch the cross-current: $94 crude is a direct cost input for miners.", None),
    ("Tankers", "NAT 6.41 / STNG 78.71", "BOTH RED at session lows while crude rips. The clearest divergence on the board. Neither is above its trigger.", RED),
    ("", "", "", None),
    ("Tonight", "TSLA, GOOGL, IBM", "Tesla options imply ~5.6% by Friday; deliveries already known (480k, +25%), so the debate is margins. Alphabet consensus $1.85 on $84.2B.", None),
    ("Tomorrow AM", "HON reports", "Which is exactly why HON is blocked today — see Why Not.", None),
    ("Today", "EIA crude, Cook 12pm ET", "No top-tier macro. With crude driving everything, the EIA inventory number is the only scheduled catalyst that matters.", None),
]
r = 4
for what, level, meaning, fill in DASH:
    row(ws, r, [what, level, meaning], fill=fill, bold_first=True)
    r += 1

# ======================================================= 3. RECOMMENDATIONS
ws = sheet("Recommendations", [10, 16, 11, 11, 11, 10, 11, 13, 62])
banner(ws, "RECOMMENDATIONS", 9, "$5,000 risk per trade | $100,000 max order | structure stop with 1x ATR floor | 2:1 minimum to nearest resistance")
header(ws, 3, ["Ticker", "Action", "Entry", "Stop", "Target", "Shares", "Notional", "Risk $", "Why this trade"])
RECS = [
    ("MU", "BUY — limit", 972.00, 891.00, 1135.00, 61, "$59,292", "$4,941",
     "The only name that clears all seven gates. Chandelier flipped LONG on 7/21 — a same-day signal, dated from the "
     "indicator's own plot series, not guessed. Price is far above ZLSMA 838. Magical reads -22.8, meaning it ran "
     "+18% over three sessions WITHOUT becoming overbought on the oscillator — that combination is rare and is the "
     "core of the case. Memory is the AI bottleneck: DRAM contract prices +93-98% quarter-over-quarter with another "
     "+58-63% projected, HBM4 shipping to NVIDIA since March. Analyst targets sit well above spot (BofA $1,550, "
     "KeyBanc $1,750, 45-analyst consensus $1,492). Stop is 1x ATR (81.38) because the 5-day structure low at 804 is "
     "17% away — too wide to be a real invalidation. Target 1135 is 2.0R and sits BELOW the 15-day high of 1168, so "
     "the 2:1 is measured to real resistance, not to open sky. Earnings are fiscal Q4, historically late September — "
     "unconfirmed but far outside this window."),
]
r = 4
for rec in RECS:
    row(ws, r, list(rec), fill=GREEN, bold_first=True)
    ws.row_dimensions[r].height = 150
    r += 1

r += 1
ws.cell(row=r, column=1, value="RISK NOTE").font = BOLD
r += 1
for line in [
    "Total committed risk: $4,941 — one position, inside the $10,000 weekly breaker with room to spare.",
    "MU is up ~18% in three sessions. That is the main argument against it: you are buying strength, not a pullback. "
    "The counter is that the oscillator says it is not yet overbought, and the signal is one day old. If you want to "
    "reduce the chase risk, half size at 30 shares and add on a pullback toward 930 keeps the thesis with less exposure.",
    "Nothing else is recommended today. That is the finding, not a gap in the work — see Why Not for each name.",
]:
    c = ws.cell(row=r, column=1, value=line)
    c.font, c.alignment = BODY, TOP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.row_dimensions[r].height = 30
    r += 1

# ============================================================== 4. WHY NOT
ws = sheet("Why Not", [10, 22, 13, 60, 44])
banner(ws, "WHY NOT — every name that did not make it", 5,
       "The specific number that killed it, and what would have to change. Sorted by how close each came.")
header(ws, 3, ["Ticker", "Failed gate", "The number", "The thinking", "What would change my mind"])
WHYNOT = [
    ("HON", "Earnings gate", "Reports 7/23 BMO",
     "This one hurts — it had the freshest Chandelier signal in the entire 76-name scan and cleared every technical "
     "gate. It reports tomorrow before the open, so a position today is a coin flip on a print, not a trade on a "
     "setup. Worse, this is the FIRST report as standalone Honeywell Technologies after the Aerospace spin-off "
     "completed 6/29 plus a 1-for-2 reverse split. Every historical level, ATR and comparable is discontinuous. "
     "Two independent reasons to stand aside.",
     "Clean print tomorrow plus a hold above the post-earnings level. Re-derive ATR and structure from post-spin "
     "bars only — pre-June levels are meaningless now.", RED),
    ("FLEX", "Earnings gate", "Reports 7/29 BMO (7 days)",
     "Passed every technical gate: Chandelier long as of 7/21, price above ZLSMA 106, Magical -68 so plenty of room. "
     "The problem is arithmetic, not thesis — the 2:1 target needs roughly 15% and earnings land in seven days. The "
     "trade cannot reach its target before the print, so taking it means holding through earnings, which the gate "
     "forbids.",
     "Either enter after 7/29 with a clean reaction, or take it deliberately as a pre-earnings momentum trade at "
     "half size with a hard exit on 7/28 — a different trade with different rules, not this one.", RED),
    ("GDX", "2:1 risk/reward", "1.0R to resistance",
     "Genuinely frustrating, because the setup is right and the macro backdrop is the best on the board — gold "
     "$4,150, miners levering it 2.5x, twelve straight quarters of earnings growth, valuations at five-year lows. "
     "But it closed +4.1% today at 77.23 and the 15-day high is 80.18. That is 2.95 of room against a 2.93 stop: "
     "exactly 1.0R. Buying here is paying for the move that already happened. The cross-current also deserves a "
     "mention — $94 crude is a direct cost input for miners, so oil strength is not unambiguously good for GDX.",
     "A pullback toward 74.50 restores the geometry to roughly 2R. Or a decisive break above 80.18 on volume, which "
     "makes the prior high support and opens clean air.", AMBER),
    ("NAT", "Trigger not hit + divergence", "6.41 vs 6.55 trigger",
     "The thesis is intact and the fundamentals are screaming — Hormuz transits down 90%, VLCC rates near $370k/day, "
     "Brent +2.8%. The stock is RED, down 0.93%, at its session low, below the 6.45 playbook trigger. When the "
     "fundamental news is this good and the equity will not go up, the equity is telling you something the news is "
     "not. Ceasefire odds also argue for a clock: 20% by 7/31 but 53% by 8/31.",
     "A close above 6.55 (yesterday's high) on real volume. Below 6.18 the setup is simply wrong. And the ceasefire "
     "kill-switch stands — this is a time-boxed trade, not a hold.", AMBER),
    ("STNG", "Trigger not hit + earnings", "78.71 vs 80.80; reports 7/30",
     "Same divergence as NAT and worse on timing — down 1.19%, at the session low, well under the 80.80 trigger, "
     "with confirmed earnings on 7/30. Even if it triggered tomorrow the print sits inside the window.",
     "Above 80.80 AND a plan that exits before 7/30, or wait for the print.", AMBER),
    ("TGTX", "Chandelier SELL mode", "Sell label 57.03",
     "Was the #1 plan before the indicator gate ran, at a clean 2.3R. The Chandelier is in SELL mode and price sits "
     "below ZLSMA 59.65. Rule 10 is absolute. Worth internalising why: the origination scan finds names in a "
     "pullback zone, but a pullback zone and a downtrend look identical on price alone — the indicator is what "
     "separates them.",
     "A Chandelier flip to long plus a reclaim of ZLSMA 59.65.", RED),
    ("EXTR", "Chandelier SELL mode", "Sell label 35.09",
     "Same story — was a 2.4R plan on zone geometry, vetoed by the indicator. Price below ZLSMA 30.75.",
     "Flip to long and a close above ZLSMA.", RED),
    ("TBLA", "Chandelier SELL mode", "Sell label 5.61",
     "Vetoed, and it was already the weakest of the three: it fell 4% INTO the bottom of its zone and sat on its "
     "10-day low. Falling to a level is not the same as bouncing off one.",
     "Flip to long, and evidence of demand at the level rather than a slide through it.", RED),
    ("SION", "Too hot + below ZLSMA", "Magical 139.5",
     "Chandelier is long but the signal dates to 6/22 — 20 sessions old, not a fresh ignition. Magical at 139.5 is "
     "pullback-only territory and price is under ZLSMA 47.71. Highest ATR in the set at 7.9%, so a wide stop for a "
     "stale signal.",
     "Cool below +100 while holding above ZLSMA, or a decisive break above the 51.57 high into open sky.", RED),
    ("CTRE", "Too hot + whipsaw", "Magical 176.1",
     "The hottest oscillator reading in the candidate set. It is also a whipsaw: sold 7/15, bought back 7/16, one bar "
     "apart. Chandelier CE(1,2) whipsaws are a known failure mode and a same-week round trip is the signature.",
     "A pullback toward ZLSMA 41.30 with the oscillator back under +100, and no further flip-flopping.", RED),
    ("PM", "Already reported + faded", "Beat, gave back ~6 pts",
     "Reported this morning: revenue $11.19B +10.4%, adjusted EPS $2.20 vs $2.04 consensus, first quarter ever above "
     "$11B. Spiked to 199.78 and sold back to ~193.60. A beat that fades on the day is a distribution tell, and the "
     "Chandelier buy from 7/16 already carries a whipsaw (sold 7/14, bought 7/16).",
     "Hold above the pre-earnings level for a few sessions. A beat that cannot hold its gap is not a base.", AMBER),
    ("CB", "Earnings + SELL mode", "Beat, round-tripped to lows",
     "Core operating $7.26 vs $6.78 estimate, combined ratio 83.8% — a clean beat. Gapped up and round-tripped to "
     "session lows. Chandelier is in SELL mode with price below ZLSMA 362.50. Note North America commercial premiums "
     "-2%, the soft spot under a good headline.",
     "Chandelier flip plus a reclaim of ZLSMA. The fundamentals are fine; the trend is not.", RED),
    ("PFE / KRE / CNOB / SKK", "Fresh but failed elsewhere", "See note",
     "All four had genuinely fresh signals (1-4 sessions). PFE Magical 113.6, too hot. KRE below ZLSMA 78.79. CNOB "
     "below ZLSMA 33.91 by half a point. SKK Magical 233.2 on 46K shares — the hottest and thinnest on the board. "
     "Freshness alone was never the bar.",
     "KRE and CNOB simply need to reclaim their ZLSMA lines — both are close. PFE and SKK need to cool off.", AMBER),
    ("VCYT / LQDA / CYRX / BRKR / OSCR / TVTX", "Extended past the zone", "+2% to +7.6% above zone tops",
     "All were valid buy-zone entries on Monday's data and all ran away on Tuesday. This is the single clearest "
     "consequence of the scanner running a day stale: by the time the list was written, the entries were gone. "
     "Buying them now is chasing a level that already paid.",
     "A pullback back INTO the stated zone with the Chandelier still long. CYRX is the closest and the only one of "
     "the six that passes the indicator stack.", AMBER),
    ("FFIV / ALKS / NWPX / VCYT / FRT / COCO / ESQ", "Earnings inside window", "7/23 through 7/31",
     "Seven names blocked purely by dates — and the scanner had missed most of these dates entirely, reporting them "
     "as unknown. Several were otherwise clean setups. FFIV in particular was genuinely in-zone.",
     "Re-run each after its print. These go back in the pool, not the bin.", None),
    ("ATLC / HELE / HOMB", "Failed 2:1", "1.8R / 1.4R / 0.3R",
     "Structure, not story. ATLC is in its zone but only 1.8R to resistance at 107.73. HELE is 1.4R and downtrending. "
     "HOMB is 0.3R, pinned right under its 15-day high. Each would need the target to clear resistance that is "
     "already overhead.",
     "HOMB becomes a breakout trigger above 31.18. ATLC needs a deeper pullback to restore the ratio.", None),
]
r = 4
for tick, gate, num, thinking, change, fill in WHYNOT:
    row(ws, r, [tick, gate, num, thinking, change], fill=fill, bold_first=True)
    ws.row_dimensions[r].height = 96
    r += 1

# ============================================================ 5. WATCH LIST
ws = sheet("Watch List", [10, 14, 20, 74])
banner(ws, "WATCH LIST — one condition away", 4, "Specific triggers. If the condition hits, the name re-enters the pool.")
header(ws, 3, ["Ticker", "Now", "Trigger", "Note"])
WATCH = [
    ("NAT", "6.41", "Close > 6.55", "Tanker thesis with a clock. Ceasefire odds 20% by 7/31 but 53% by 8/31 — a 3-6 week window, time-boxed. Kill-switch on any ceasefire headline."),
    ("STNG", "78.71", "Close > 80.80", "Same thesis, but earnings 7/30 sit inside the window. Needs a plan that exits before the print."),
    ("GDX", "77.23", "Pullback to ~74.50", "Best macro backdrop on the board; only the entry price is wrong. A pullback restores roughly 2R."),
    ("CYRX", "16.24", "Pullback to 15.12-15.73", "The one origination name that passes the full indicator stack. Just needs to come back into its zone."),
    ("HOMB", "30.42", "Break > 31.18", "Fails 2:1 at current price; becomes a clean breakout trigger above the 15-day high."),
    ("KRE", "75.98", "Reclaim ZLSMA 78.79", "Fresh signal 7/16, only the trend line is missing. Regional banks tie to the rate-relief thesis."),
    ("CNOB", "33.41", "Reclaim ZLSMA 33.91", "Fresh signal 7/15, half a point below the line."),
    ("HON", "237.19", "After 7/23 print", "Freshest signal in the scan. Re-derive all levels from post-spin bars only."),
    ("FLEX", "128.33", "After 7/29 print", "Clean technicals, blocked only by the calendar."),
]
r = 4
for w in WATCH:
    row(ws, r, list(w), bold_first=True)
    ws.row_dimensions[r].height = 34
    r += 1

# ============================================================= 6. SELL MODE
ws = sheet("Sell Mode", [14, 90])
banner(ws, "SELL MODE — rule 10 veto list", 2, "Chandelier is short on these. Do not buy, regardless of the story.")
header(ws, 3, ["Count", "Tickers"])
SELL = ["TGTX", "EXTR", "TBLA", "CB", "FFIV", "NWPX", "LQDA", "BRKR", "OSCR", "ATLC", "HELE", "PNC", "GD", "GOOG",
        "GS", "HD", "INTC", "JNJ", "KO", "LAES", "MCD", "MDT", "MP", "NEM", "NFLX", "NVAX", "NVTS", "PEP", "RIVN",
        "UAVS", "HUT", "HIMX", "MCHP"]
row(ws, 4, [f"{len(SELL)} names", ", ".join(SELL)])
ws.row_dimensions[4].height = 60
ws.cell(row=6, column=1, value="Note").font = BOLD
c = ws.cell(row=6, column=2, value="GOOG is on this list and reports tonight. INTC is on it despite closing +8.6% "
                                   "yesterday — a big up day inside a downtrend is not a trend change. That is the "
                                   "whole point of the veto.")
c.font, c.alignment = BODY, TOP
ws.row_dimensions[6].height = 34

# ====================================================== 7. METHOD & QUALITY
ws = sheet("Method & Data Quality", [104])
banner(ws, "METHOD & DATA QUALITY", 1, "Read this before trusting any number above.")
r = 3
for text, font in [
    ("HOW SIGNALS WERE DERIVED", BOLD),
    ("Every name was read through the O.G Chandelier stack — Chandelier Exit + ZLSMA + Magical OB/OS — on the live "
     "daily chart, layout 152878770. Note that the layout previously recorded for this scan (91698392) does NOT "
     "carry the Magical study; that has been corrected.", BODY),
    ("Signal dates are EXACT, not estimated. data_get_pine_labels returns nothing because the Chandelier draws with "
     "plotshape rather than label.new — but the study's own 400-bar plot series is readable directly from the chart "
     "data model. That is the indicator's real output, not a reimplementation.", BODY),
    ("", BODY),
    ("CORRECTION LOGGED", BOLD),
    ("An earlier pass inferred signal age from how far the buy label sat below current price. That was WRONG — the "
     "label prints at the Chandelier STOP level, not an entry price. It made MU's same-day flip look 27% stale. All "
     "dates here come from the plot series instead.", BODY),
    ("", BODY),
    ("KNOWN GAPS — these are real", BOLD),
    ("1. The 7/21 origination scan was computed on 7/20 closes. Verified across all 29 names: every scanner price "
     "equalled that name's prior-day close. Seven names it flagged as in-zone had already run out of them. The "
     "scanner needs a same-day-bar assertion before its list is trusted again.", BODY),
    ("2. Rotation radar and ignition sweep have NOT run — python execution was blocked. No radar deltas, no "
     "market-wide ignition pass. Both are still owed.", BODY),
    ("3. The fundamental gate is off — no Finviz file, so this is a technical-only screen. No float, EPS or sales "
     "screening was applied to anything here.", BODY),
    ("4. Watchlist coverage is PARTIAL. The watchlist reader returns only rendered rows (54, starting at GD), so "
     "names alphabetically before GD were never scanned. 76 names total were covered.", BODY),
    ("5. MU and NAT earnings dates are UNVERIFIED. Both were left null rather than promoted from conflicting "
     "aggregator estimates. MU's pattern is late September; NAT's is late August. Treat as earnings-exposed then.", BODY),
    ("6. The system remains UNPROVEN. 143 names in the leader log and 394 in the recommendation tracker, and NOT ONE "
     "has a matured 21-day result yet. There is no measured track record behind any of this. Size accordingly.", BODY),
    ("", BODY),
    ("SOURCING", BOLD),
    ("Prices: live TradingView feed only. Web price quotes were contradictory today (three different values for the "
     "same crude print) and were discarded in favour of the feed. News: Reuters, Bloomberg, CNBC, SEC filings, "
     "company IR and primary newswires only. Anything unverifiable was labelled rather than dropped or guessed.", BODY),
]:
    c = ws.cell(row=r, column=1, value=text)
    c.font, c.alignment = font, TOP
    r += 1

out_dir = os.path.join(REPO, "reports")
os.makedirs(out_dir, exist_ok=True)
out = os.path.join(out_dir, f"curated_report_{DATE}.xlsx")
wb.save(out)
print("written", out)
