"""SHADOW detector: WOC (Washed-Out Coil).

STATUS: SHADOW LANE ONLY. Pre-registered 2026-07-22. NOT a live rule.
Governance: rule freeze through 2026-08-31 (commit 2d673c6). Promotion requires
20-30 out-of-sample signals beating the incumbent. Until then this only logs.

WHY IT EXISTS
-------------
2026-07-21: SMCI was scanned by the origination screen and dumped on the
"Excluded - not screenable" tab (score 27.2, 38.2% of 52-wk high). The next
session it opened +4% and ran +34% intraday. Both live screens are trend-
CONTINUATION / leader screens: they structurally cannot see a washed-out name
turning. The Excluded tab is where that shape goes to die, ~900 rows a day.

This detector reads the tab the live system throws away and asks a single
question: is this a broken chart that has stopped going down and coiled?

It is deliberately coarse - 5 conditions, no weights, no score. Anything more
tuned would be fitting to one stock (SMCI) and would not survive out-of-sample.

USAGE
    python scripts/shadow_woc_detector.py                 # all workbooks found
    python scripts/shadow_woc_detector.py 2026-07-21      # one date
"""
import glob
import json
import os
import re
import sys
from datetime import datetime

import openpyxl

# ---------------------------------------------------------------------------
# PRE-REGISTERED RULE - frozen 2026-07-22. Do not tune without a new git commit
# and a new pre-registration entry, or the out-of-sample count resets to zero.
# ---------------------------------------------------------------------------
RULE = {
    'version': 'woc-v1',
    'pct_of_52wk_high_max': 50.0,   # washed out
    'coiled_spring': 'Y',           # range has compressed
    'relvol_today_max': 1.5,        # BEFORE the ignition bar - a "before photo"
    'rsi_daily_min': 25.0,          # not in freefall
    'rsi_daily_max': 50.0,          # not already recovered
    'adx_min': 20.0,                # some trend structure exists
}

COLS = {
    'ticker': 'Ticker',
    'price': 'Price',
    'pct52': '% of 52-Wk High',
    'adr': 'Avg Daily Move %',
    'relvol': 'Volume vs Normal (today)',
    'relvol3': 'Volume vs Normal (3-day)',
    'heavy': 'Heavy-Volume Days (last 5)',
    'rsi': 'RSI Daily (heat gauge)',
    'rsiw': 'RSI Weekly',
    'adx': 'Trend Strength (ADX)',
    'adx_rising': 'Trend Getting Stronger?',
    'coiled': 'Coiled Spring?',
    'above50': '% Above 50-Day Avg',
    'stop': 'Suggested Stop ($)',
}


def _num(v):
    """Coerce '38.2%' / '-28.0%' / 25 / None -> float or None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('%', '').replace(',', '')
    if not s or s in {'-', 'N/A', 'None'}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def scan_workbook(path):
    """Return (date, [hit dicts], rows_examined) for one origination workbook."""
    m = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(path))
    date = m.group(1) if m else '?'

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if 'Excluded' not in wb.sheetnames:
        return date, [], 0
    ws = wb['Excluded']

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return date, [], 0
    header = [str(c).strip() if c is not None else '' for c in rows[0]]
    idx = {key: header.index(name) for key, name in COLS.items() if name in header}

    hits, examined = [], 0
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        examined += 1
        g = lambda k: row[idx[k]] if k in idx and idx[k] < len(row) else None

        pct52 = _num(g('pct52'))
        relvol = _num(g('relvol'))
        rsi = _num(g('rsi'))
        adx = _num(g('adx'))
        coiled = str(g('coiled') or '').strip().upper()

        if pct52 is None or relvol is None or rsi is None or adx is None:
            continue
        if pct52 > RULE['pct_of_52wk_high_max']:
            continue
        if coiled != RULE['coiled_spring']:
            continue
        if relvol > RULE['relvol_today_max']:
            continue
        if not (RULE['rsi_daily_min'] <= rsi <= RULE['rsi_daily_max']):
            continue
        if adx < RULE['adx_min']:
            continue

        hits.append({
            'ticker': str(g('ticker')).strip(),
            'price': _num(g('price')),
            'pct_of_52wk_high': pct52,
            'rsi_daily': rsi,
            'rsi_weekly': _num(g('rsiw')),
            'adx': adx,
            'adx_rising': str(g('adx_rising') or '').strip(),
            'relvol_today': relvol,
            'relvol_3day': _num(g('relvol3')),
            'heavy_vol_days': _num(g('heavy')),
            'adr_pct': _num(g('adr')),
            'pct_above_50d': _num(g('above50')),
            'suggested_stop': _num(g('stop')),
        })

    hits.sort(key=lambda h: (-(h['adx'] or 0), h['pct_of_52wk_high']))
    return date, hits, examined


def main():
    want = sys.argv[1] if len(sys.argv) > 1 else None
    paths = sorted(glob.glob('reports/origination_scan_*.xlsx'))
    if want:
        paths = [p for p in paths if want in p]
    if not paths:
        print('no origination_scan_*.xlsx found in reports/')
        return

    stamp = datetime.now().strftime('%Y-%m-%d %H:%M local')
    print(f'SHADOW WOC DETECTOR  [{RULE["version"]}]   written at {stamp}')
    print('STATUS: shadow lane - logging only, NOT a live rule\n')
    print(f'  rule: <={RULE["pct_of_52wk_high_max"]:.0f}% of 52wk high, coiled=Y, '
          f'relvol<{RULE["relvol_today_max"]}, RSI {RULE["rsi_daily_min"]:.0f}-'
          f'{RULE["rsi_daily_max"]:.0f}, ADX>={RULE["adx_min"]:.0f}\n')

    ledger = {'rule': RULE, 'written_at': stamp, 'days': {}}
    for path in paths:
        date, hits, examined = scan_workbook(path)
        ledger['days'][date] = {'examined': examined, 'count': len(hits), 'hits': hits}
        names = ', '.join(h['ticker'] for h in hits) if hits else '(none)'
        print(f'{date}   {len(hits):>3} hits / {examined} excluded rows')
        print(f'          {names}\n')

    out = 'research/shadow-woc-ledger.json'
    os.makedirs('research', exist_ok=True)
    with open(out, 'w') as fh:
        json.dump(ledger, fh, indent=1)
    print(f'wrote {out}')

    counts = [d['count'] for d in ledger['days'].values()]
    if counts:
        print(f'\nSIGNAL RATE: min {min(counts)}, max {max(counts)}, '
              f'mean {sum(counts)/len(counts):.1f} per day')
        print('A detector firing on dozens of names a day is noise, not signal.')


if __name__ == '__main__':
    main()
