"""Daily auto-feed for Omar's Crude_Fundamental_Dashboard workbook.

Pulls the freshest hard data and writes it into an "AUTO FEED" sheet in the
newest Downloads copy of the workbook - WITHOUT touching any other tab
(Inputs stays hand-curated; a timestamped backup is made before every save).

Sources (per Omar 2026-08-05: lean on TradingView/TradeStation, not web text):
  - WTI + Brent futures curve (front 4 contracts each) -> TradingView's own
    futures scanner API (same feed as the desktop app)
  - Cushing / SPR / crude / gasoline / distillate stocks, production,
    refinery utilization, exports -> EIA weekly history pages (the primary
    source itself; new numbers land Wednesdays 9:30 CT)

Run daily (part of the EOD routine):
  python scripts/update_crude_dashboard_feed.py            # newest workbook
  python scripts/update_crude_dashboard_feed.py <path.xlsx>
"""
import datetime as dt
import glob
import json
import os
import re
import shutil
import sys
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font

DOWNLOADS = Path.home() / "Downloads"
REPO = Path(__file__).resolve().parent.parent
QUOTED_JSON = REPO / "research" / "crude-quoted-latest.json"
BBL_PER_TONNE_GASOIL = 7.45  # ICE LS Gasoil is $/tonne; convert to $/bbl
SCAN_URL = "https://scanner.tradingview.com/futures/scan"
EIA_URL = "https://www.eia.gov/dnav/pet/hist/LeafHandler.ashx?n=PET&s={sid}&f=W"
MONTH_CODES = "FGHJKMNQUVXZ"  # Jan..Dec

EIA_SERIES = [
    ("Cushing crude stocks (kbbl)", "W_EPC0_SAX_YCUOK_MBBL", "Cushing inventory (kbbl)"),
    ("SPR crude stocks (kbbl)", "WCSSTUS1", "US SPR (kbbl)"),
    ("US commercial crude (kbbl)", "WCESTUS1", "US crude C+C row, Inventory Tracker"),
    ("US total gasoline (kbbl)", "WGTSTUS1", "US gasoline row, Inventory Tracker"),
    ("US distillate (kbbl)", "WDISTUS1", "US distillate row, Inventory Tracker"),
    ("US crude production (kbpd)", "WCRFPUS2", "US production (MMbpd) /1000"),
    ("US refinery utilization (%)", "WPULEUS3", "US refinery util (%)"),
    ("US crude exports (kbpd)", "WCREXUS2", "US crude exports (MMbpd) /1000"),
]


def tv_scan(tickers):
    r = requests.post(SCAN_URL, json={"symbols": {"tickers": tickers},
                                      "columns": ["name", "close", "change"]}, timeout=30)
    r.raise_for_status()
    return {row["s"]: row["d"] for row in r.json().get("data", [])}


def candidate_contracts(prefix, exchange, n_months=10):
    today = dt.date.today()
    out = []
    for i in range(n_months):
        m = today.month + i
        y = today.year + (m - 1) // 12
        m = (m - 1) % 12 + 1
        out.append(f"{exchange}:{prefix}{MONTH_CODES[m - 1]}{y}")
    return out


def get_curve(prefix, exchange, front_cont):
    """Return [(contract, close, day_change_pct)] for the front 4 live contracts."""
    cands = candidate_contracts(prefix, exchange)
    data = tv_scan(cands + [f"{exchange}:{front_cont}"])
    front_close = data.get(f"{exchange}:{front_cont}", [None, None])[1]
    live = [(t, data[t][1], data[t][2]) for t in cands if t in data and data[t][1]]
    # anchor the strip at the contract matching the continuous front price
    start = 0
    if front_close is not None:
        for i, (_, close, _) in enumerate(live):
            if abs(close - front_close) < 0.02:
                start = i
                break
    return live[start:start + 4]


def eia_last_two(series_id):
    """Return [(date_str, value), (prev_date, prev_value)] newest first."""
    html = requests.get(EIA_URL.format(sid=series_id), timeout=30).text
    pairs = []
    year_month = None
    for m in re.finditer(r"class='B6'>\s*(?:&nbsp;)*([\d]{4})-(\w{3})|class='B5'>([\d/]{4,5})&nbsp;|class='B3'>([\d,\.]+)&nbsp;", html):
        if m.group(1):
            year_month = m.group(1)
        elif m.group(3):
            pairs.append([f"{m.group(3)}/{year_month}" if year_month else m.group(3), None])
        elif m.group(4) and pairs and pairs[-1][1] is None:
            pairs[-1][1] = float(m.group(4).replace(",", ""))
    pairs = [p for p in pairs if p[1] is not None]
    return pairs[-1:] + pairs[-2:-1]  # newest, prior


def month_label(contract):
    code = contract.split(":")[1]
    m = re.match(r"[A-Z]+([FGHJKMNQUVXZ])(\d{4})", code)
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{months[MONTH_CODES.index(m.group(1))]}{m.group(2)[2:]}" if m else code


def tv_to_ts(contract):
    """NYMEX:CLU2026 -> CLU26 (TradeStation futures format)."""
    code = contract.split(":")[1]
    m = re.match(r"([A-Z]+)([FGHJKMNQUVXZ])(\d{4})", code)
    return f"{m.group(1)}{m.group(2)}{m.group(3)[2:]}" if m else code


def ts_layer(wti):
    """TradeStation cross-check / fallback for the US-listed legs.
    Returns (note_line, wti_curve_or_None). Never raises."""
    try:
        sys.path.insert(0, str(Path(__file__).parent))
        import ts_api
        if not ts_api.available():
            return ("TradeStation: key stored, awaiting one-time login (scripts/ts_auth.py)", None)
        if wti:  # cross-check the front contract
            sym = tv_to_ts(wti[0][0])
            q = ts_api.ts_quotes([sym]).get(sym)
            if q:
                diff = abs(float(q["Last"]) - wti[0][1])
                return (f"TradeStation cross-check {sym}: {q['Last']} vs TV {wti[0][1]} "
                        f"(diff {diff:.2f}) {'OK' if diff < 0.25 else 'DIVERGENT - investigate'}", None)
            return ("TradeStation cross-check: no quote returned", None)
        # TV curve failed -> rebuild WTI strip from TS
        cands = candidate_contracts("CL", "NYMEX")
        ts_syms = [tv_to_ts(c) for c in cands]
        quotes = ts_api.ts_quotes(ts_syms)
        curve = [(cands[i], float(quotes[s]["Last"]), 0.0)
                 for i, s in enumerate(ts_syms) if s in quotes][:4]
        return ("TradeStation FALLBACK supplied the WTI curve (TV scan empty)", curve or None)
    except Exception as e:
        return (f"TradeStation layer unavailable: {e}", None)


def newest_workbook():
    cands = [p for p in glob.glob(str(DOWNLOADS / "Crude_Fundamental_Dashboard*.xlsx"))
             if not os.path.basename(p).startswith("~$")]
    if not cands:
        sys.exit("No Crude_Fundamental_Dashboard*.xlsx in Downloads")
    return Path(max(cands, key=os.path.getmtime))


def main():
    wb_path = Path(sys.argv[1]) if len(sys.argv) > 1 else newest_workbook()
    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M")

    wti = get_curve("CL", "NYMEX", "CL1!")
    brn = get_curve("BRN", "ICEEUR", "BRN1!")
    ts_note, ts_curve = ts_layer(wti)
    if not wti and ts_curve:
        wti = ts_curve

    # TS is Omar's REAL-TIME entitled NYMEX feed; the public TV scanner is
    # ~10-min delayed on CME (proven 8/5: 0.32 stable gap on a fast tape).
    # So TS overrides the US legs when authenticated; TV keeps ICE legs.
    ts_rb = ts_ho = None
    try:
        sys.path.insert(0, str(Path(__file__).parent))
        import ts_api
        if ts_api.available() and wti:
            syms = [tv_to_ts(t) for t, _, _ in wti[:4]]
            mcode = re.match(r"CL([FGHJKMNQUVXZ]\d{2})", syms[0]).group(1)
            got = ts_api.ts_quotes(syms + [f"RB{mcode}", f"HO{mcode}"])
            wti = [(t, float(got[tv_to_ts(t)]["Last"]), chg) if tv_to_ts(t) in got
                   else (t, c, chg) for t, c, chg in wti]
            if f"RB{mcode}" in got:
                ts_rb = float(got[f"RB{mcode}"]["Last"])
            if f"HO{mcode}" in got:
                ts_ho = float(got[f"HO{mcode}"]["Last"])
            ts_note = f"US legs = TradeStation REAL-TIME (WTI strip + RB/HO {mcode}); TV = ICE legs + fallback"
    except Exception as e:
        ts_note = f"TS override skipped ({e}); TV values in use"

    # products + freight proxy (TS real-time overrides RB/HO when set above)
    prod = tv_scan(["NYMEX:RB1!", "NYMEX:HO1!", "ICEEUR:ULS1!"])
    rb = ts_rb or prod.get("NYMEX:RB1!", [None, None, None])[1]   # $/gal
    ho = ts_ho or prod.get("NYMEX:HO1!", [None, None, None])[1]   # $/gal
    uls = prod.get("ICEEUR:ULS1!", [None, None, None])[1]         # $/tonne (ICE - TV only)
    try:
        r = requests.post("https://scanner.tradingview.com/america/scan",
                          json={"symbols": {"tickers": ["AMEX:BWET"]},
                                "columns": ["name", "close", "change"]}, timeout=30)
        bwet = r.json()["data"][0]["d"] if r.ok and r.json().get("data") else None
    except Exception:
        bwet = None
    cl1 = wti[0][1] if wti else None
    brn1 = brn[0][1] if brn else None
    cracks = {}
    if rb and ho and cl1:
        cracks["RBOB crack (RB*42 - WTI)"] = round(rb * 42 - cl1, 2)
        cracks["ULSD crack (HO*42 - WTI)"] = round(ho * 42 - cl1, 2)
        cracks["USGC 3-2-1 proxy ((2RB+HO)*42/3 - WTI)"] = round((2 * rb + ho) * 42 / 3 - cl1, 2)
    if uls and brn1:
        cracks["NWE gasoil crack (ULS/7.45 - Brent)"] = round(uls / BBL_PER_TONNE_GASOIL - brn1, 2)

    # article-quoted physicals harvested from the morning brief (written at EOD
    # by Claude from the draft's MACHINE BLOCK) - carried with source + as-of
    quoted = []
    if QUOTED_JSON.exists():
        try:
            quoted = json.loads(QUOTED_JSON.read_text(encoding="utf-8")).get("items", [])
        except Exception as e:
            quoted = [{"key": "PARSE ERROR", "value": str(e), "source": "", "asof": ""}]

    # East-West spread proxies: TV Brent M1 minus the wrap-quoted Mideast
    # outrights (exchange APIs are gated - CME ToS-blocked, ICE Cloudflare,
    # DME unreachable - so Oman/Dubai/Murban ride the daily Reuters wrap)
    def quoted_num(key):
        for q in quoted:
            if q.get("key") == key:
                m = re.search(r"-?\d+(?:\.\d+)?", str(q.get("value", "")))
                if m:
                    return float(m.group()), q.get("asof", "")
        return None, ""

    ew_rows = []
    if brn1:
        for key, label in [("OMAN_M1", "Brent - Oman (EFS proxy via DME settle quote)"),
                           ("DUBAI_M1", "Brent - Dubai (EFS via quoted Dubai M1)"),
                           ("MURBAN_M1", "Brent - Murban (via quoted IFAD settle)")]:
            val, asof = quoted_num(key)
            if val is not None:
                ew_rows.append((label, round(brn1 - val, 2),
                                f"Brent {brn1} (TV live) - {val} (as-of {asof})"))
    eia = []
    for label, sid, maps_to in EIA_SERIES:
        try:
            rows = eia_last_two(sid)
            latest = rows[0] if rows else ("n/a", None)
            prior = rows[1] if len(rows) > 1 else ("n/a", None)
            eia.append((label, latest, prior, maps_to))
        except Exception as e:  # one bad series must not kill the run
            eia.append((label, (f"FETCH FAILED: {e}", None), ("", None), maps_to))

    # backup, then write the AUTO FEED sheet only
    bdir = DOWNLOADS / "dashboard_backups"
    bdir.mkdir(exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    shutil.copy2(wb_path, bdir / f"{wb_path.stem}-{stamp}.xlsx")

    wb = openpyxl.load_workbook(wb_path)
    if "AUTO FEED" in wb.sheetnames:
        del wb["AUTO FEED"]
    ws = wb.create_sheet("AUTO FEED", 0)
    bold = Font(bold=True)

    def put(row, col, val, is_bold=False):
        c = ws.cell(row=row, column=col, value=val)
        if is_bold:
            c.font = bold

    put(1, 1, "AUTO FEED - written by scripts/update_crude_dashboard_feed.py", True)
    put(2, 1, f"Updated: {now} (Houston). Sources: TradingView futures feed + EIA weekly + TradeStation cross-check.")
    put(2, 8, ts_note)
    put(3, 1, "This sheet is machine-owned and rewritten on every run. Reference these")
    put(4, 1, "cells from Inputs (or paste values) - nothing else in the file is touched.")

    r = 6
    put(r, 1, "FUTURES CURVE (TradingView live/settle)", True)
    r += 1
    for col, h in enumerate(["Benchmark", "M1", "M2", "M3", "M4", "M1-M2", "M2-M3", "M3-M4"], 1):
        put(r, col, h, True)
    for name, curve in (("WTI (NYMEX CL)", wti), ("Brent (ICE BRN)", brn)):
        r += 1
        put(r, 1, name)
        for i, (t, close, _) in enumerate(curve[:4]):
            put(r, 2 + i, f"{month_label(t)}  {close}")
        if len(curve) >= 4:
            closes = [c for _, c, _ in curve[:4]]
            put(r, 6, round(closes[0] - closes[1], 2))
            put(r, 7, round(closes[1] - closes[2], 2))
            put(r, 8, round(closes[2] - closes[3], 2))
    if wti and brn:
        r += 1
        put(r, 1, "WTI-Brent (M1)")
        put(r, 2, round(wti[0][1] - brn[0][1], 2))

    r += 2
    put(r, 1, "PRODUCTS & CRACKS (computed from TradingView futures - daily)", True)
    r += 1
    for col, h in enumerate(["Metric", "Value", "Maps to"], 1):
        put(r, col, h, True)
    for label, val, maps in [
        ("RBOB M1 ($/gal)", rb, ""),
        ("ULSD/HO M1 ($/gal)", ho, ""),
        ("ICE LS Gasoil M1 ($/tonne)", uls, ""),
    ]:
        r += 1
        put(r, 1, label); put(r, 2, val); put(r, 3, maps)
    for label, val in cracks.items():
        r += 1
        put(r, 1, label); put(r, 2, val)
        put(r, 3, "Crack Spreads tab market column (futures-based proxy)")
    if bwet:
        r += 1
        put(r, 1, "BWET tanker ETF (freight DIRECTION proxy, not a rate)")
        put(r, 2, f"{bwet[1]}  ({bwet[2]:+.1f}% today)")

    if ew_rows:
        r += 2
        put(r, 1, "EAST-WEST SPREADS (computed: TV Brent live minus wrap-quoted Mideast outrights)", True)
        r += 1
        for col, h in enumerate(["Spread", "Value ($/bbl)", "Legs"], 1):
            put(r, col, h, True)
        for label, val, legs in ew_rows:
            r += 1
            put(r, 1, label); put(r, 2, val); put(r, 3, legs)
            if "Oman" in label or "Dubai" in label:
                put(r, 4, "Maps to: East-West Arb tab, Brent-Dubai EFS market column")

    r += 2
    put(r, 1, "ARTICLE-QUOTED PHYSICALS (harvested from the morning brief - dated, not daily-guaranteed)", True)
    r += 1
    for col, h in enumerate(["Item", "Value", "Source", "As-of"], 1):
        put(r, col, h, True)
    if quoted:
        for q in quoted:
            r += 1
            put(r, 1, q.get("key")); put(r, 2, q.get("value"))
            put(r, 3, q.get("source")); put(r, 4, q.get("asof"))
    else:
        r += 1
        put(r, 1, "(none harvested yet - populated from the daily brief's machine block)")

    r += 2
    put(r, 1, "EIA WEEKLY (new data Wednesdays 9:30 CT)", True)
    r += 1
    for col, h in enumerate(["Series", "Latest w/e", "Value", "Prior", "WoW change", "Maps to (Inputs)"], 1):
        put(r, col, h, True)
    for label, latest, prior, maps_to in eia:
        r += 1
        put(r, 1, label)
        put(r, 2, latest[0])
        put(r, 3, latest[1])
        put(r, 4, prior[1])
        if latest[1] is not None and prior[1] is not None:
            put(r, 5, round(latest[1] - prior[1], 1))
        put(r, 6, maps_to)

    wb.save(wb_path)

    # console summary for the chat / EOD log
    print(f"AUTO FEED written to {wb_path.name} (backup in {bdir.name}/)")
    for name, curve in (("WTI", wti), ("Brent", brn)):
        strip = " | ".join(f"{month_label(t)} {c}" for t, c, _ in curve[:4])
        if len(curve) >= 4:
            closes = [c for _, c, _ in curve[:4]]
            strip += f"  ->  spreads {closes[0]-closes[1]:+.2f} / {closes[1]-closes[2]:+.2f} / {closes[2]-closes[3]:+.2f}"
        print(f"{name}: {strip}")
    for label, latest, prior, _ in eia:
        wow = ""
        if latest[1] is not None and prior[1] is not None:
            wow = f" (WoW {latest[1]-prior[1]:+,.0f})"
        print(f"{label}: {latest[1]} w/e {latest[0]}{wow}")
    for label, val in cracks.items():
        print(f"{label}: {val}")
    for label, val, legs in ew_rows:
        print(f"{label}: {val}  [{legs}]")
    if bwet:
        print(f"BWET freight proxy: {bwet[1]} ({bwet[2]:+.1f}% today)")
    print(f"Quoted physicals carried: {len(quoted)}" + ("" if quoted else " (run the EOD harvest to populate)"))
    print(ts_note)


if __name__ == "__main__":
    main()
