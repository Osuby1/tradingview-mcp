"""Shared Excel formatting standard for ALL scan workbooks (long and short).

Set 2026-07-22 after Omar rejected the layout of reports/universe_2026-07-21.xlsx.
The origination scan workbook is the reference for look-and-feel; this module
copies what that file does well and fixes what BOTH files did badly.

WHAT THE ORIGINATION SCAN DOES RIGHT (copied here)
--------------------------------------------------
* A "READ ME FIRST" tab that states the reading order explicitly.
* Sheet names numbered in that reading order: "1 - BUY ZONE", "2 - ...".
* Navy header band (1F3A5F), white bold text, wrapped, centered, frozen.
* **A hover tooltip on EVERY column header** written in plain English, ending
  with a pointer back to the READ ME tab. The universe workbook had 30 such
  tooltips on the origination side and ZERO on its own - that is the single
  biggest readability gap between the two files.
* Plain-English column titles with the jargon in parentheses:
  "Trend Strength (ADX)", not "ADX".

WHAT BOTH FILES DID BADLY (fixed here)
--------------------------------------
1. **Notes as an unreadable blob.** universe_2026-07-21 crammed 200+ character
   narratives into a 46-wide "Notes" column. Unreadable in Excel, so nobody
   reads it. Fixed by splitting into WHY (one short line, on the data row) and
   a separate "Notes & Decisions" sheet with wrapped text and tall rows.

2. **Data-quality findings buried as repeated rows.** The 7/21 Summary sheet had
   eleven rows all keyed "Data quality", each a wall of prose. One of them said:

       "Watchlist coverage is PARTIAL: watchlist_get returns only rendered DOM
        rows (54 of them, starting at GD), so names alphabetically before GD
        were not scanned."

   That is the exact bug that made SMCI invisible - it was already KNOWN and
   WRITTEN DOWN on 7/21, and it still got missed on 7/22, because it was the
   ninth identical-looking row in a wall of prose. Fixed by a dedicated
   "Data Quality" sheet with a SEVERITY column, colour-coded, BLOCKERS sorted
   to the top, plus an explicit "What To Do" column.

3. **Mojibake.** Em-dashes and middle dots rendered as garbage. All text written
   through `safe()` which keeps output ASCII-clean.

USAGE
-----
    from scan_workbook_style import Col, write_table, readme_sheet, \
        data_quality_sheet, notes_sheet, verdict_fill

    cols = [Col("sym", "Ticker", 9, "The stock symbol."),
            Col("last", "Price", 10, "Latest price.", fmt=FMT_PRICE)]
    write_table(ws, cols, rows)
"""
import math

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- theme ------------------------------------------------------------------
# Header band: GREEN with black bold text. Omar restyled universe_2026-07-27 by
# hand (green band, black text, boxed READ ME sections, auto-fit row heights)
# and on 2026-07-28 made that the standard for every daily scan workbook. The
# navy band below survives only as the SECTION font colour.
NAVY = "1F3A5F"          # section-heading colour (kept from the old theme)
HDR_GREEN = "92D050"     # header band, per Omar's 2026-07-27 hand-edit
HDR_FONT = Font(bold=True)
HDR_FILL = PatternFill("solid", fgColor=HDR_GREEN)
HDR_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")

GREEN = PatternFill("solid", fgColor="C6EFCE")   # actionable / pass
AMBER = PatternFill("solid", fgColor="FFF2CC")   # conflicted / caution
RED = PatternFill("solid", fgColor="FFC7CE")     # blocked / blocker
GREY = PatternFill("solid", fgColor="EDEDED")    # informational

BOLD = Font(bold=True)
TITLE = Font(bold=True, size=14)
SECTION = Font(bold=True, size=11, color=NAVY)
RED_FONT = Font(bold=True, color="C00000")

TITLE_FILL = PatternFill("solid", fgColor=HDR_GREEN)   # READ ME banner
_MEDIUM = Side(style="medium")

# number formats
FMT_PRICE = "#,##0.00"
FMT_PCT = "0.0"
FMT_INT = "#,##0"
FMT_X = "0.00"

READ_ME_POINTER = "\n\n(Full guide: 'READ ME FIRST' tab)"

_SUBS = {
    # real unicode punctuation
    "—": " - ", "–": "-", "·": "|", "’": "'", "‘": "'",
    "“": '"', "”": '"', "→": "->", "≥": ">=", "≤": "<=", "×": "x",
    # UTF-8 bytes previously misdecoded as cp1252 ("mojibake"). The upstream
    # JSON already contains these, so fixing only real unicode is not enough -
    # that is why "closes — verified" rendered as 'closes ??" verified'.
    "â€”": " - ", "â€“": "-", "â€™": "'", "â€˜": "'",
    "â€œ": '"', "â€\x9d": '"', "â€¢": "*", "Â·": "|", "Â": "",
    "�": "",
}


def unmojibake(v):
    """Repair UTF-8 bytes that were previously decoded as cp1252.

    The upstream universe-results JSON already contains these (an em-dash shows
    up as the three characters U+00E2 U+20AC U+201D), so substituting individual
    literals is guesswork - round-tripping through cp1252 fixes the whole class
    at once. Returns the input unchanged if the round-trip is not valid UTF-8.
    """
    if not isinstance(v, str) or "â" not in v and "Â" not in v:
        return v
    try:
        repaired = v.encode("cp1252", "strict").decode("utf-8", "strict")
        return repaired
    except (UnicodeEncodeError, UnicodeDecodeError):
        return v


def safe(v):
    """Keep cell text ASCII-clean, repairing mojibake first."""
    if not isinstance(v, str):
        return v
    v = unmojibake(v)
    for bad, good in _SUBS.items():
        v = v.replace(bad, good)
    return v.encode("ascii", "replace").decode("ascii")


def plainify(v):
    """Translate leftover indicator shorthand into Omar's plain English.

    Applied to long-text cells (why / reasoning / notes). New text should be
    written plain at the source; this catches strings that already live in old
    universe-results JSON files, so recompiling an old date reads right too.
    Each substitution is guarded so already-plain text is not doubled up.
    """
    if not isinstance(v, str) or not v:
        return v
    t = v
    if "ZLSMA" in t and "trend line" not in t:
        t = t.replace("above ZLSMA", "above its trend line (ZLSMA)")
        t = t.replace("below ZLSMA", "below its trend line (ZLSMA)")
        t = t.replace("the ZLSMA", "the trend line (ZLSMA)")
        if "trend line" not in t:
            t = t.replace("ZLSMA", "trend line (ZLSMA)", 1)
    if "ADX" in t and "trend strength" not in t and "trend-strength" not in t:
        t = t.replace("ADX", "trend strength (ADX)", 1)
    if "x ATR" in t and "day's swing" not in t and "daily noise" not in t:
        t = t.replace("x ATR", "x ATR (multiples of a normal day's swing)", 1)
    if ("+DI" in t or "-DI" in t) and "sellers" not in t and "buyers" not in t:
        t = t.replace("+DI > -DI", "buyers ahead of sellers (+DI > -DI)")
    if "vs 200-day)" in t:
        t = t.replace("vs 200-day)", "vs the 200-day average)")
    return t


def autofit_rows(ws, cols, first_row=2, last_row=None, min_h=15, max_h=90):
    """Approximate Excel's auto-fit so wrapped text is fully visible.

    Omar's 2026-07-27 hand-edit auto-fit every data tab; without it a wrapped
    cell shows one line and the rest is invisible until you click it. openpyxl
    cannot measure text, so estimate: lines = chars / usable column width,
    15pt per line, capped so one monster cell cannot make a 300pt row.
    """
    last_row = last_row or ws.max_row
    for r in range(first_row, last_row + 1):
        lines = 1
        for i, c in enumerate(cols, 1):
            val = ws.cell(row=r, column=i).value
            if val is None:
                continue
            width = max(6.0, (c.width or 10) - 2)
            need = 0
            for seg in str(val).split("\n"):
                need += max(1, math.ceil(len(seg) / width))
            lines = max(lines, need)
        ws.row_dimensions[r].height = min(max_h, max(min_h, 15 * lines))


def box_region(ws, r1, r2, c1=1, c2=2):
    """Medium border box around a block - the READ ME section framing Omar added."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            b = cell.border
            cell.border = Border(
                left=_MEDIUM if c == c1 else b.left,
                right=_MEDIUM if c == c2 else b.right,
                top=_MEDIUM if r == r1 else b.top,
                bottom=_MEDIUM if r == r2 else b.bottom)


_GRID = Border(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM)


def grid_borders(ws, r1, r2, c1, c2):
    """Medium border around EVERY cell in the block - each row reads as a boxed
    card. Omar hand-drew this grid on the Plans tab of universe_2026-07-27_new
    and made it the standard for that tab (2026-07-28)."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = _GRID


def side_table(ws, cols, rows, start_col, row_fill=None, grid=True):
    """A second table rendered to the RIGHT of a sheet's main table.

    Added 2026-07-28 for Omar's actioned-trades block on the Track Record tab:
    same header styling and tooltips as write_table, but anchored at start_col
    instead of column A, with the Plans-style border grid so it reads as its
    own card. Returns the last row written.
    """
    for i, c in enumerate(cols):
        col_n = start_col + i
        cell = ws.cell(row=1, column=col_n, value=safe(c.title))
        cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
        if c.tip:
            com = Comment(safe(c.tip) + READ_ME_POINTER, "scan")
            com.width, com.height = 340, 190
            cell.comment = com
        ws.column_dimensions[get_column_letter(col_n)].width = c.width
    for rn, r in enumerate(rows, 2):
        fill = row_fill(r) if row_fill else None
        for i, c in enumerate(cols):
            cell = ws.cell(row=rn, column=start_col + i, value=safe(r.get(c.key)))
            if c.fmt and isinstance(cell.value, (int, float)):
                cell.number_format = c.fmt
            cell.alignment = Alignment(
                wrap_text=True, vertical="top",
                horizontal=c.align or ("left" if i == 0 else "center"))
            if fill:
                cell.fill = fill
    if grid:
        grid_borders(ws, 1, 1 + len(rows), start_col, start_col + len(cols) - 1)
    return 1 + len(rows)


class Col:
    """One column: storage key, display title, width, and the hover tooltip.

    `tip` is REQUIRED. A column with no tooltip is a column the reader has to
    guess at, which is what made the old universe workbook unusable.
    """

    __slots__ = ("key", "title", "width", "tip", "fmt", "align")

    def __init__(self, key, title, width, tip, fmt=None, align=None):
        if not tip:
            raise ValueError(f"column {title!r} needs a plain-English tooltip")
        self.key, self.title, self.width = key, title, width
        self.tip, self.fmt, self.align = tip, fmt, align


def style_header_row(ws, titles, tips=None, widths=None, freeze="A2",
                     autofilter=True):
    """Apply the standard navy header band, tooltips, widths, freeze, filter."""
    for i, title in enumerate(titles, 1):
        c = ws.cell(row=1, column=i, value=safe(title))
        c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
        if tips and i - 1 < len(tips) and tips[i - 1]:
            com = Comment(safe(tips[i - 1]) + READ_ME_POINTER, "scan")
            com.width, com.height = 340, 190
            c.comment = com
        if widths and i - 1 < len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[1].height = 34
    if freeze:
        ws.freeze_panes = freeze
    if autofilter and titles:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(titles))}1"


def write_table(ws, cols, rows, freeze="A2", autofilter=True, row_fill=None,
                autofit=True):
    """Write a full table from Col specs + a list of dicts.

    row_fill: optional fn(row_dict) -> PatternFill or None, applied across the row.
    autofit: size data rows to their wrapped text (Omar's 2026-07-27 standard).
             Sheets that want a fixed height set it after this call, which wins.
    """
    style_header_row(ws, [c.title for c in cols], [c.tip for c in cols],
                     [c.width for c in cols], freeze, autofilter)
    for r in rows:
        vals = [safe(r.get(c.key)) for c in cols]
        ws.append(vals)
        n = ws.max_row
        fill = row_fill(r) if row_fill else None
        for i, c in enumerate(cols, 1):
            cell = ws.cell(row=n, column=i)
            if c.fmt and isinstance(cell.value, (int, float)):
                cell.number_format = c.fmt
            cell.alignment = Alignment(
                wrap_text=True, vertical="top",
                horizontal=c.align or ("left" if i <= 2 else "center"))
            if fill:
                cell.fill = fill
    if autofit and rows:
        autofit_rows(ws, cols)
    return ws


def verdict_fill(row):
    """Standard colour rule for a Verdict column, used by every scan."""
    v = str(row.get("verdict") or "").upper()
    if v.startswith(("ACTIVE", "NEW ACTIVE", "TRIGGER")) and "CONFLICT" not in v:
        return GREEN
    if "CONFLICT" in v or "HALF" in v or "WATCH" in v:
        return AMBER
    if v.startswith(("BLOCKED", "VETO", "SKIP")):
        return RED
    return None


def readme_sheet(ws, title, stamps, what_this_is, sheet_guide, key_readings,
                 how_to_act=None):
    """Structured READ ME FIRST - headed sections, not a wall of prose.

    sheet_guide: list of (sheet_name, one_line_purpose, optional_second_line)
    """
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 96

    # Green title banner + medium-border boxes around each section: Omar's
    # 2026-07-27 hand-edit, now the standard for every daily workbook.
    ws.append([safe(title)])
    ws["A1"].font = TITLE
    ws["A1"].fill = TITLE_FILL
    ws["B1"].fill = TITLE_FILL
    ws.row_dimensions[1].height = 19
    ws.append([])
    _start = ws.max_row + 1
    for label, value in stamps:
        ws.append([safe(label), safe(value)])
        ws.cell(row=ws.max_row, column=1).font = BOLD
    box_region(ws, _start, ws.max_row)
    ws.append([])

    _start = ws.max_row + 1
    ws.append(["WHAT THIS IS"])
    ws.cell(row=ws.max_row, column=1).font = SECTION
    for line in what_this_is:
        ws.append(["", safe(line)])
    box_region(ws, _start, ws.max_row)
    ws.append([])

    _start = ws.max_row + 1
    ws.append(["READ THE TABS IN THIS ORDER"])
    ws.cell(row=ws.max_row, column=1).font = SECTION
    for entry in sheet_guide:
        name, purpose = entry[0], entry[1]
        ws.append([safe(name), safe(purpose)])
        ws.cell(row=ws.max_row, column=1).font = BOLD
        for extra in entry[2:]:
            ws.append(["", safe("    " + extra)])
    box_region(ws, _start, ws.max_row)
    ws.append([])

    if how_to_act:
        _start = ws.max_row + 1
        ws.append(["HOW TO ACT ON IT"])
        ws.cell(row=ws.max_row, column=1).font = SECTION
        for line in how_to_act:
            ws.append(["", safe(line)])
        box_region(ws, _start, ws.max_row)
        ws.append([])

    _start = ws.max_row + 1
    ws.append(["WHAT THE NUMBERS MEAN"])
    ws.cell(row=ws.max_row, column=1).font = SECTION
    for label, meaning in key_readings:
        ws.append([safe(label), safe(meaning)])
        ws.cell(row=ws.max_row, column=1).font = BOLD
    box_region(ws, _start, ws.max_row)

    for row in ws.iter_rows(min_row=1):
        for c in row:
            if c.column == 2:
                c.alignment = Alignment(wrap_text=True, vertical="top")
    return ws


SEVERITY_ORDER = {"BLOCKER": 0, "WARNING": 1, "INFO": 2}
SEVERITY_FILL = {"BLOCKER": RED, "WARNING": AMBER, "INFO": GREY}


def data_quality_sheet(ws, items):
    """Severity-ranked data-quality sheet.

    items: list of dicts with severity / area / finding / impact / action.
    Accepts bare strings for backwards compatibility (treated as INFO).

    This exists because on 2026-07-21 the watchlist truncation bug WAS recorded
    and still got missed - it was the ninth undifferentiated 'Data quality' row
    in a wall of prose. Blockers now sort to the top and are coloured red.
    """
    cols = [
        Col("severity", "Severity", 12,
            "BLOCKER = this run's conclusions may be WRONG until fixed. "
            "WARNING = a real limitation, read before acting. INFO = context."),
        Col("area", "Area", 20, "Which part of the pipeline the issue is in."),
        Col("finding", "What Went Wrong", 60,
            "The problem, stated plainly in one or two sentences."),
        Col("impact", "Effect On This Run", 52,
            "What this actually means for the numbers in this workbook - "
            "which conclusions you can and cannot trust."),
        Col("action", "What To Do", 42,
            "The concrete fix or the check to run before trusting this again."),
    ]
    norm = []
    for it in items:
        if isinstance(it, str):
            it = {"severity": "INFO", "area": "", "finding": it,
                  "impact": "", "action": ""}
        it.setdefault("severity", "INFO")
        norm.append(it)
    norm.sort(key=lambda x: SEVERITY_ORDER.get(str(x["severity"]).upper(), 3))
    write_table(ws, cols, norm,
                row_fill=lambda r: SEVERITY_FILL.get(str(r["severity"]).upper()))
    for n in range(2, ws.max_row + 1):
        ws.row_dimensions[n].height = 58
    return ws


def notes_sheet(ws, rows):
    """Per-ticker narrative, given room to breathe instead of a 46-wide cell.

    rows: dicts with sym / verdict / changed / reasoning.
    """
    cols = [
        Col("rank", "Rank", 7,
            "Buy-score rank from the Fresh Buys tab, so the two tabs read in the "
            "same order."),
        Col("sym", "Ticker", 10, "The stock symbol."),
        Col("verdict", "Verdict", 24,
            "What the frozen rule set says to do with this name."),
        Col("changed", "What Changed", 34,
            "How this call differs from the previous run: NEW (was not a fresh "
            "signal last time), UPGRADED, DOWNGRADED, or same call as before. "
            "Compared automatically against the prior results file."),
        Col("pros", "PROS - what argues FOR it", 54,
            "Every gate this name passed and how strongly. Derived from the same "
            "components as the buy score, so it cannot drift from the ranking."),
        Col("cons", "CONS - what argues AGAINST it", 54,
            "Every gate it failed or only marginally passed. If this column is "
            "empty the name is genuinely clean; if it is long, the rank is "
            "flattering it."),
        Col("reasoning", "Full Reasoning", 80,
            "The complete argument: which gates passed, which failed, and the "
            "judgement call. This is the detail that used to be crushed into "
            "the Notes column on the data tabs."),
    ]
    write_table(ws, cols, rows, row_fill=verdict_fill)
    for n in range(2, ws.max_row + 1):
        ws.row_dimensions[n].height = 92
    return ws


def split_note(note):
    """Split a legacy blob note into (what_changed, why_one_line, full_text)."""
    if not note:
        return "", "", ""
    note = safe(note).strip()
    changed = ""
    for tag in ("REVERSAL OF EARLIER CALL.", "REVERSAL.", "DOWNGRADED but coherent.",
                "DOWNGRADED.", "UPGRADED.", "NEW."):
        if note.startswith(tag):
            changed = tag.rstrip(".")
            note = note[len(tag):].strip()
            break
    first = note.split(". ")[0].strip()
    if len(first) > 110:
        first = first[:107] + "..."
    return changed, first, note
