#!/usr/bin/env python3
"""Compile the day's run-the-universe results into a formatted Excel workbook.

Reads watchlists/universe-results-<date>.json (written by the run-the-universe
command) and writes reports/universe_<date>.xlsx with four sheets:
  Fresh Buys & Plans | Blocked | Sell Mode | Summary

Freshness rule: only the requested date (default today) is compiled. If that
file is missing - e.g. the universe leg died on credits - the compiler exits 2
rather than silently republishing an older run, so a failed scan can never
masquerade as a fresh report. Pass --allow-stale to compile the newest file
anyway; the workbook is then stamped STALE on the Summary sheet.
"""
import argparse
import datetime
import glob
import json
import os
import sys

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def results_for(date):
    path = os.path.join(REPO, "watchlists", f"universe-results-{date}.json")
    return path if os.path.exists(path) else None


def newest_results():
    files = sorted(glob.glob(os.path.join(REPO, "watchlists", "universe-results-*.json")))
    return files[-1] if files else None


def main():
    ap = argparse.ArgumentParser(description="Compile run-the-universe results into Excel.")
    ap.add_argument("--date", default=datetime.date.today().isoformat(),
                    help="scan date to compile, YYYY-MM-DD (default: today)")
    ap.add_argument("--allow-stale", action="store_true",
                    help="compile the newest results file even if it is older than --date")
    args = ap.parse_args()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        sys.exit("openpyxl not installed - pip install openpyxl")

    path = results_for(args.date)
    requested = None
    if not path:
        newest = newest_results()
        if not args.allow_stale:
            have = f"newest is {os.path.basename(newest)}" if newest else "none exist at all"
            print(f"SKIPPED: no universe-results-{args.date}.json ({have}). The universe leg did not "
                  f"write results for {args.date}; refusing to republish an older report. "
                  f"Re-run the universe leg, or pass --allow-stale to compile the older file anyway.",
                  file=sys.stderr)
            sys.exit(2)
        if not newest:
            print("SKIPPED: --allow-stale given but no universe-results-*.json exists at all.",
                  file=sys.stderr)
            sys.exit(2)
        path = newest
        requested = args.date
        print(f"WARNING: no results for {requested} - compiling STALE {os.path.basename(path)}.")

    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    date = data.get("date") or os.path.basename(path)[17:27] or datetime.date.today().isoformat()

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F4E79")

    def sheet(title, headers, rows, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title
        ws.append(headers)
        for c in ws[1]:
            c.font, c.fill = head_font, head_fill
        for r in rows:
            ws.append(r)
        for i, h in enumerate(headers, 1):
            width = max([len(str(h))] + [len(str(r[i - 1])) for r in rows if len(r) >= i]) + 2
            ws.column_dimensions[ws.cell(1, i).column_letter].width = min(width, 60)
        ws.freeze_panes = "A2"
        return ws

    hits = data.get("hits", [])
    sheet(
        "Fresh Buys & Plans",
        ["Ticker", "Source", "Signal date", "Last price", "Entry", "Dist %", "Stop",
         "Size", "Risk $", "Targets", "Earnings gate", "Magical", "Regime", "Verdict", "Notes"],
        [[h.get("sym"), h.get("source", "watchlist"), h.get("fresh"), h.get("last"),
          h.get("entry"), h.get("dist_pct"), h.get("stop"), h.get("size"), h.get("risk"),
          h.get("targets"), h.get("earnings"), h.get("magical"), h.get("regime"),
          h.get("verdict"), h.get("note")] for h in hits],
        first=True,
    )
    sheet(
        "Blocked",
        ["Ticker", "Source", "Signal date", "Reason"],
        [[b.get("sym"), b.get("source", "watchlist"), b.get("fresh"), b.get("reason")]
         for b in data.get("blocked", [])],
    )
    sm = data.get("sell_mode", [])
    sheet("Sell Mode (veto)", ["Ticker"], [[s] for s in sm])
    sheet(
        "Summary",
        ["Field", "Value"],
        [["Scan date", date],
         ["Data as of", date],
         ["Written at", datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")],
         ["Freshness", f"STALE - {requested} was requested but the newest results were {date}"
                       if requested else "fresh"],
         ["Variant", data.get("variant", "?")],
         ["Universe size", data.get("universe_size", "?")],
         ["Fresh buys", len(hits)],
         ["Active plans", sum(1 for h in hits if str(h.get("verdict", "")).lower().startswith("active"))],
         ["Blocked", len(data.get("blocked", []))],
         ["Sell mode", len(sm)],
         ["Combined plan risk $", data.get("combined_risk", "?")],
         ["Notes", data.get("notes", "")]],
    )

    out_dir = os.path.join(REPO, "reports")
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.join(out_dir, f"universe_{date}.xlsx")
    wb.save(out)
    print(f"Report written: {out}")


if __name__ == "__main__":
    main()
