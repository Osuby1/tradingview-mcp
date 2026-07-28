#!/usr/bin/env python3
"""Export the origination scan's actionable tabs into the repo for "run the universe".

Reads Documents/Equities_Scanner/origination_scan.xlsx (the stage2_leader_scanner_v3
output) and writes watchlists/origination-tabs.md with the tickers from the
Buy Zone / Fresh Ignitions / Coiled tabs, so the run-the-universe command can
merge them into the O.G Chandelier scan universe.

Run right after the daily scan:
    python stage2_leader_scanner_v3.py && python <repo>/scripts/export_origination_tabs.py
or standalone any time after the xlsx exists. Robust to tab-name and ticker-column
naming variations; skips missing tabs with a warning instead of failing.
"""
import datetime

# US market holidays that matter for the current window. Extend as needed - a
# missing holiday only shifts the archive name by one day, it does not corrupt data.
_HOLIDAYS_2026 = {
    "2026-01-01", "2026-01-19", "2026-02-16", "2026-04-03", "2026-05-25",
    "2026-06-19", "2026-07-03", "2026-09-07", "2026-11-26", "2026-12-25",
}


def data_date_from_workbook(xlsx_path):
    """Read the scanner's own 'DATA AS OF: <date>' stamp. Authoritative.

    Added 2026-07-22 alongside the scanner change that writes it. Prefer this
    over any clock-based guess: Yahoo's posting lag varies, so inferring the
    session from the run time is unreliable. A scan run at 15:50 on 2026-07-21
    - well after the close - still returned 07-20 data.

    Returns an ISO date string, or None if the stamp is absent (older workbook).
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if "READ ME FIRST" not in wb.sheetnames:
            return None
        for row in wb["READ ME FIRST"].iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str) and "DATA AS OF:" in cell:
                    m = re.search(r"(\d{4}-\d{2}-\d{2})", cell)
                    if m:
                        return m.group(1)
        return None
    except Exception:
        return None


def data_session_date(run_dt):
    """Which trading session does a scan started at `run_dt` actually contain?

    The scanner reads the most recent COMPLETED daily bar. So:
      * run AFTER today's close  -> today's session
      * run BEFORE today's close -> the previous trading day
      * run on a weekend/holiday -> the previous trading day

    This is why origination_scan_2026-07-21.xlsx held 07-20 prices: it was
    produced by a scan that ran while the 7/21 session was still open. Naming
    the archive by the RUN date made the file lie about its own contents, and
    every downstream join was silently one session off.

    `run_dt` is local (America/Chicago on this machine); the close is 15:00 local.
    """
    CLOSE_HOUR_LOCAL = 15  # 3pm Central = 4pm Eastern

    d = run_dt.date()
    if run_dt.hour < CLOSE_HOUR_LOCAL:
        d -= datetime.timedelta(days=1)
    # walk back over weekends and holidays
    for _ in range(10):
        if d.weekday() < 5 and d.isoformat() not in _HOLIDAYS_2026:
            return d.isoformat()
        d -= datetime.timedelta(days=1)
    return d.isoformat()
import json
import os
import re
import sys

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCAN_DIR = os.path.join(os.path.expanduser("~"), "Documents", "Equities_Scanner")
OUT = os.path.join(REPO, "watchlists", "origination-tabs.md")


def newest_scan_xlsx():
    """The freshest scanner output, whatever filename it landed on.

    BUG FOUND 2026-07-28: this script read the fixed name origination_scan.xlsx
    only. On 7/27 Omar had that file open in Excel, so the scanner - which had
    fresh Monday data - saved to its fallback name origination_scan_152744.xlsx
    instead. The export silently read the stale locked file and Monday's
    workbook joined FRIDAY's origination scores. Pick the newest matching file
    by modified time so the fallback name is found automatically.
    """
    import glob
    cands = [c for c in glob.glob(os.path.join(SCAN_DIR, "origination_scan*.xlsx"))
             if not os.path.basename(c).startswith("~$")]   # skip Excel lock stubs
    return max(cands, key=os.path.getmtime) if cands else None

# case-insensitive substring match per target tab
TAB_PATTERNS = {"Buy Zone": "buy", "Fresh Ignitions": "ignit", "Coiled": "coil"}
# tracked past recommendations - exported separately (exit-watch, NOT merged into the buy-scan universe)
TRACKER_PATTERN = "track"
TICKER_RE = re.compile(r"^[A-Z][A-Z0-9.\-]{0,6}$")


def sheet_tickers(ws):
    """Pull tickers from a worksheet: prefer a Ticker/Symbol column, else column A."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    col = 0
    for i, h in enumerate(header):
        if "ticker" in h or "symbol" in h:
            col = i
            break
    out = []
    for row in rows[1:]:
        if col >= len(row) or row[col] is None:
            continue
        val = str(row[col]).strip().upper()
        val = val.split(":")[-1]  # strip any EXCHANGE: prefix
        if TICKER_RE.match(val) and val not in out:
            out.append(val)
    return out


def main():
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl not installed — pip install openpyxl")
    XLSX = newest_scan_xlsx()
    if not XLSX:
        sys.exit(f"No origination_scan*.xlsx in {SCAN_DIR} — run the origination scan first.")
    if os.path.basename(XLSX) != "origination_scan.xlsx":
        print(f"NOTE: using fallback-named scan file {os.path.basename(XLSX)} "
              f"(origination_scan.xlsx was locked when the scanner saved) - it is "
              f"the newest output on disk.")

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    mtime = datetime.datetime.fromtimestamp(os.path.getmtime(XLSX))
    tabs, missing = {}, []
    for label, pat in TAB_PATTERNS.items():
        match = next((n for n in wb.sheetnames if pat in n.lower()), None)
        if match is None:
            missing.append(label)
            continue
        tabs[label] = sheet_tickers(wb[match])

    # Tracker source = recommendations_log.csv (clean header) - the xlsx TRACKER
    # tab has a decorative layout that defeats column sniffing (learned 7/19).
    tracker, tracker_note = [], ""
    reco_csv = os.path.join(os.path.dirname(XLSX), "recommendations_log.csv")
    if os.path.exists(reco_csv):
        import csv
        with open(reco_csv, newline="", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            fields = reader.fieldnames or []
            low = [c.lower() if c else "" for c in fields]
            tick_col = next((fields[i] for i, c in enumerate(low) if "ticker" in c or "symbol" in c), None)
            date_col = next((fields[i] for i, c in enumerate(low) if "date" in c or c == "day"), None)
            price_col = next((fields[i] for i, c in enumerate(low)
                              if "price" in c or "entry" in c or "close" in c or c.endswith("px") or "ref" in c), None)
            if tick_col:
                seen = set()
                for row in reader:
                    val = str(row.get(tick_col) or "").strip().upper().split(":")[-1]
                    if not TICKER_RE.match(val) or val in seen:
                        continue  # keep FIRST occurrence = earliest recommendation
                    seen.add(val)
                    rec = {"sym": val}
                    if date_col:
                        rec["rec_date"] = str(row.get(date_col) or "").strip()
                    if price_col:
                        try:
                            rec["rec_price"] = float(str(row.get(price_col)).replace("$", "").replace(",", ""))
                        except (TypeError, ValueError):
                            pass
                    tracker.append(rec)
                if not (date_col and price_col):
                    tracker_note = f"(date/price columns not fully identified - headers: {', '.join(map(str, fields))})"
            else:
                tracker_note = f"(no ticker column found in recommendations_log.csv - headers: {', '.join(map(str, fields))})"
    else:
        tracker_note = "(recommendations_log.csv not found)"

    all_syms = sorted({s for syms in tabs.values() for s in syms})
    # Prefer the scanner's own stamp; fall back to the clock-based guess only
    # for workbooks written before that stamp existed.
    data_date_str = data_date_from_workbook(XLSX)
    data_date_src = "scanner stamp"
    if not data_date_str:
        data_date_str = data_session_date(mtime)
        data_date_src = "INFERRED from run time - scanner predates the DATA AS OF stamp"
    _expected = data_session_date(datetime.datetime.now())
    if data_date_str < _expected:
        print(f"WARNING: origination data is STALE - file carries {data_date_str} "
              f"data but the last completed session is {_expected}. Every "
              f"downstream join will be a session behind until the scan re-runs.")
    lines = [
        f"# Origination scan tabs — exported {datetime.date.today().isoformat()}",
        "",
        f"Source: {os.path.basename(XLSX)} (modified {mtime:%Y-%m-%d %H:%M}). "
        "Feeds the run-the-universe command: these tickers are merged into the "
        "O.G Chandelier scan universe alongside the three watchlists.",
        "",
        f"**DATA AS OF: {data_date_str} close** ({data_date_src}). The scanner reads "
        f"the most recent COMPLETED session at run time, so a scan run intraday "
        f"carries the PREVIOUS day's data. Join on this date, not on the export date.",
        "",
    ]
    for label in TAB_PATTERNS:
        syms = tabs.get(label, [])
        lines.append(f"## {label} ({len(syms)})")
        lines.append(", ".join(syms) if syms else "(tab not found)" if label in missing else "(empty)")
        lines.append("")
    lines.append(f"## Tracker — logged recommendations, exit-watch only ({len(tracker)})")
    lines.append(", ".join(r["sym"] for r in tracker) if tracker else (tracker_note or "(empty)"))
    if tracker_note and tracker:
        lines.append(tracker_note)
    lines.append("")
    lines += [
        "```json",
        json.dumps(
            {"exported": datetime.date.today().isoformat(),
             "data_as_of": data_date_str,
             "xlsx_modified": mtime.strftime("%Y-%m-%d %H:%M"),
             "tabs": tabs, "all": all_syms, "tracker": tracker},
            indent=1),
        "```",
        "",
    ]
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    # dated archive of the scanner's xlsx (it overwrites itself daily).
    # Named for the DATA date, not the run date - see data_session_date().
    import shutil
    arch_dir = os.path.join(REPO, "reports")
    os.makedirs(arch_dir, exist_ok=True)
    data_date = data_date_str
    arch = os.path.join(arch_dir, f"origination_scan_{data_date}.xlsx")
    try:
        shutil.copy2(XLSX, arch)
        print(f"Archived: {arch}")
        if data_date != datetime.date.today().isoformat():
            print(f"  NOTE: scan ran {mtime:%Y-%m-%d %H:%M} local, so its data is "
                  f"through {data_date} - archived under the DATA date, not today.")
    except OSError as e:
        print(f"WARNING: dated xlsx archive failed: {e}")

    print(f"Exported {len(all_syms)} unique tickers "
          f"({', '.join(f'{k}: {len(v)}' for k, v in tabs.items())})")
    print(f"Tracker (exit-watch): {len(tracker)} tickers {tracker_note}".rstrip())
    if missing:
        print(f"WARNING — tabs not found in xlsx: {', '.join(missing)} "
              f"(available sheets: {', '.join(wb.sheetnames)})")
    print(f"Written: {OUT}")


if __name__ == "__main__":
    main()
