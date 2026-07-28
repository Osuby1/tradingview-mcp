#!/usr/bin/env python3
"""Compile the run-the-universe results into a rich Excel workbook (v3).

v3 (2026-07-22): rebuilt on scripts/scan_workbook_style.py after Omar rejected
the v2 layout. Every column now carries a plain-English hover tooltip, tabs are
numbered in reading order, long notes moved out of a 46-wide cell onto their own
sheet, and data-quality findings are severity-ranked instead of being eleven
identical-looking rows of prose.

Reads:
  - watchlists/universe-results-<date>.json  (all_names + scanner_enrich; hits/verdicts)
  - reports/origination_scan_<date>.xlsx     (Score/Grade/Category join, optional)
  - watchlists/rotation-radar-state.md       (radar states, optional)
  - watchlists/universe-context-<date>.json  (market context lines, optional)

Usage: python scripts/compile_universe_report_v2.py [YYYY-MM-DD]
Writes: reports/universe_<date>.xlsx
"""
import datetime
import json
import os
import re
import sys

from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from scan_workbook_style import (  # noqa: E402
    Col, write_table, readme_sheet, data_quality_sheet, notes_sheet,
    verdict_fill, split_note, safe, style_header_row, plainify, grid_borders,
    side_table,
    FMT_PRICE, FMT_PCT, FMT_X, FMT_INT, BOLD, RED_FONT, GREEN, AMBER, RED, GREY, SECTION,
)
from gate_stack import evaluate, funnel, MissingGateData, DEFAULT_POSITION  # noqa: E402
from rank_candidates import (  # noqa: E402
    score_long, score_long_v2, score_short, rank_blocked_severity,
    SHORT_POLICY, RANK_RATIONALE,
)

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATE = sys.argv[1] if len(sys.argv) > 1 else None
# Optional second arg: explicit output path. Lets a re-compile be checked
# without touching a workbook Omar may have hand-annotated.
OUT_OVERRIDE = sys.argv[2] if len(sys.argv) > 2 else None
if not DATE:
    cands = sorted(f for f in os.listdir(os.path.join(REPO, "watchlists"))
                   if re.match(r"universe-results-\d{4}-\d{2}-\d{2}\.json$", f))
    DATE = cands[-1][17:-5]

import schema  # noqa: E402

_RES_PATH = os.path.join(REPO, "watchlists", f"universe-results-{DATE}.json")
res = json.load(open(_RES_PATH))
# The workbook is the deliverable Omar acts on, so it must never be built from a
# file shape it does not understand. Legacy files still compile (with a note).
schema.require(res, understood=(2, 3), path=_RES_PATH)
enrich = res.get("scanner_enrich", {})
allrows = {r["sym"]: r for r in res.get("all_names", [])}
hits = {h["sym"]: h for h in res.get("hits", [])}

# --- origination join (Score/Grade/Category) ---
orig = {}
opath = os.path.join(REPO, "reports", f"origination_scan_{DATE}.xlsx")
if os.path.exists(opath):
    wb0 = load_workbook(opath, read_only=True)
    for ws0 in wb0.worksheets:
        if not re.match(r"\d ", ws0.title):
            continue
        rows0 = ws0.iter_rows(values_only=True)
        header = [str(c) if c else "" for c in next(rows0)]
        try:
            it, isc, igr, icat = (header.index(x) for x in
                                  ("Ticker", "Score (0-100)", "Grade", "Category / What To Do"))
        except ValueError:
            continue
        for row in rows0:
            if row[it] and str(row[it]).strip() not in orig:
                orig[str(row[it]).strip()] = {"score": row[isc], "grade": row[igr],
                                              "cat": row[icat], "tab": ws0.title}

# --- radar parse ---
# BUG FOUND 2026-07-28: this only ever understood a {"groups":[{group,state,accel,
# perf_w,perf_1m,perf_6m}]} shape. rotation_radar.py has never written that - it
# writes {"date":..., "states": {"AMEX:XLE": "IGNITING", ...}}, a ticker->state map
# and nothing else. So `groups` was always [], radar_lines was always empty, and
# the workbook printed "(rotation radar did not run)" on every single run while the
# radar was in fact running fine. Classic silent schema drift.
#
# Fix: read the states map as well. The numbers it lacks (accel / 1W / 1M / 6M) are
# already computed by market_structure.py against the SAME universe, so the table is
# joined from the two at render time. Old shape still honoured if it ever appears.
radar_lines = []
radar_states = {}
rpath = os.path.join(REPO, "watchlists", "rotation-radar-state.md")
if os.path.exists(rpath):
    txt = open(rpath, encoding="utf-8").read()
    m = re.search(r"```json\s*(.*?)```", txt, re.S)
    if m:
        try:
            rj = json.loads(m.group(1))
            if isinstance(rj, dict) and isinstance(rj.get("states"), dict):
                radar_states = rj["states"]
            groups = rj.get("groups", rj if isinstance(rj, list) else [])
            for g in sorted(groups, key=lambda x: -(x.get("accel") or 0)):
                radar_lines.append({"group": g.get("group", g.get("name", "?")),
                                    "state": g.get("state", "?"), "accel": g.get("accel"),
                                    "w": g.get("perf_w"), "m1": g.get("perf_1m"),
                                    "m6": g.get("perf_6m")})
        except Exception:
            pass

ctx = {}
cpath = os.path.join(REPO, "watchlists", f"universe-context-{DATE}.json")
if os.path.exists(cpath):
    ctx = json.load(open(cpath))

WRITTEN = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
ENRICH_AS_OF = res.get("scanner_enrich_asof", "n/a")
wb = Workbook()


def plan_display(row, ha_last, ha_stop):
    """The REAL, tradeable (price, stop, risk%) for anything Omar reads.

    Added 2026-07-25. Signals stay on Heikin Ashi; plans must not. When the sweep
    captured real prices (schema v3+) the stop is carried across at the SAME
    PERCENTAGE distance, which preserves the risk geometry the Chandelier
    expressed while quoting levels that exist in the market. Legacy files with no
    real price fall back to the HA numbers, so old workbooks still compile.
    """
    entry = row.get("plan_entry") or row.get("real_close") or row.get("real_last")
    if not entry:
        risk = (round((ha_last - ha_stop) / ha_last * 100, 1)
                if (ha_last and ha_stop and ha_last > ha_stop) else None)
        return ha_last, ha_stop, risk
    stop = row.get("plan_stop")
    if stop is None and ha_last and ha_stop and ha_last > ha_stop > 0:
        stop = round(entry * (1 - (ha_last - ha_stop) / ha_last), 4)
    risk = (round((entry - stop) / entry * 100, 1)
            if (stop and entry > stop) else None)
    return round(entry, 4), stop, risk


def _fmt_cap(x):
    """Money in human units: 19770000000 -> '$19.8B'."""
    if not x:
        return None
    for suf, div in (("B", 1e9), ("M", 1e6), ("K", 1e3)):
        if abs(x) >= div:
            return f"${x/div:.1f}{suf}"
    return f"${x:.0f}"


def classify(note):
    """Legacy data-quality strings are plain prose. Rank them so the dangerous
    ones cannot hide. The 7/21 watchlist-truncation bug was recorded and still
    missed because every note looked equally important."""
    u = note.upper()
    if any(k in u for k in ("PARTIAL", "NOT SCANNED", "WAS WRONG", "STALE",
                            "CORRECTION", "ARE VOID", "OFF-STANDARD")):
        sev, act = "BLOCKER", "Fix before trusting the coverage or numbers in this run."
    elif any(k in u for k in ("SKIPPED", "DID NOT RUN", "REPAINT", "VERIFY",
                              "ARTIFACT", "APPROXIMATE", "NOT TRADEABLE")):
        sev, act = "WARNING", "Confirm on the live chart before sizing anything."
    else:
        sev, act = "INFO", ""
    area = ("Coverage" if "WATCHLIST" in u or "SCANNED" in u else
            "Data feed" if any(k in u for k in ("CHART", "PRICE", "HEIKIN", "SPLIT")) else
            "Signal quality" if any(k in u for k in ("REPAINT", "LABEL", "SIGNAL")) else
            "Pipeline")
    return {"severity": sev, "area": area, "finding": note,
            "impact": "", "action": act}


# ---------------------------------------------------------------- READ ME ---
ws = wb.active
ws.title = "READ ME FIRST"
readme_sheet(
    ws,
    f"RUN-THE-UNIVERSE WORKBOOK - {DATE}",
    stamps=[("Scan data as of", f"{DATE} close (live chart reads)"),
            ("Scanner enrich as of", ENRICH_AS_OF),
            ("Workbook written", f"{WRITTEN} local"),
            ("Generated by", "compile_universe_report_v2.py (v3 layout)")],
    what_this_is=[
        "Every stock on the watchlists, plus the origination scanner's actionable names,",
        "read through the O.G Chandelier stack (Chandelier Exit 1/2 + ZLSMA-50 + Magical",
        "OB/OS) on the LIVE chart, on HEIKIN ASHI (Omar's 2026-07-23 standard).",
        "",
        "The origination scanner finds candidates. THIS run tells you which ones actually",
        "triggered, which ones the gates rejected, and why.",
    ],
    sheet_guide=[
        ("1 - Market & Rotation", "The tape and where money is flowing. Read this first.",
         "If the regime is hostile, everything below gets smaller size or no size."),
        ("2 - Fresh Buys", "Every buy signal 5 sessions old or newer.",
         "The Verdict column is the decision. Green = actionable, amber = conflicted, red = blocked."),
        ("2b - Gated Longs", "Every stock in BUY mode - of ANY age - that still clears the full gate stack.",
         "Fresh Buys only lists signals 5 sessions old or newer; this tab is where the",
         "established, still-valid uptrends live (the DVN/PBR type). Ranked least-stretched",
         "first, so the best entries sit on top. Check the Extension and Catalyst columns",
         "before acting - many older uptrends are stretched or about to report earnings."),
        ("3 - Plans", "The sized trades: entry, stop, share count, dollar risk, targets, alerts."),
        ("4 - Blocked", "Fresh signals the gates rejected, and which gate did it.",
         "These are NOT trades. The reason column is the point of this tab."),
        ("5 - Sell Mode", "Names the indicator has in downtrend. Rule 10: do not buy these."),
        ("6 - Tracker Broken", "Past recommendations whose trend has broken. Exit / bank-it flags."),
        ("7 - Notes & Decisions", "The full reasoning per name, with room to actually read it.",
         "Anything that reversed or was downgraded since the last run is flagged here."),
        ("8 - Market Movers", "The day's biggest gainers/losers in the WHOLE market, each screened.",
         "Almost all are microcap SKIPs; the few liquid movers get a real buy/short call."),
        ("9 - Track Record", "Every past origination pick, one row each, showing how it has done",
         "since the day it was called - price then, price now, best and worst point along",
         "the way. The grade scorecard sits below the list (so far a higher grade has NOT",
         "been worth more). To the RIGHT: the MY TRADES block - the recommendations Omar",
         "actually put money on (open + closed), with real fills and a running tally of",
         "banked vs paper profit."),
        ("10 - Data Quality", "What was broken or partial in this run, ranked by severity.",
         "READ THE RED ROWS. A BLOCKER means conclusions in this file may be wrong."),
        ("11 - Run Summary", "Counts and run metadata."),
    ],
    how_to_act=[
        "1. Check tab 10 for BLOCKERS first. If coverage was partial, absence of a name",
        "   proves nothing - it may simply never have been scanned.",
        "2. Read the regime on tab 1. It sets your size for everything else.",
        "3. Work tab 2 top-down. Only green Verdicts are candidates.",
        "4. Cross-check each candidate against tab 5 - a sell-mode name is never a buy.",
        "5. Size it on tab 3. If the stop is wider than the risk unit allows, skip it.",
    ],
    key_readings=[
        ("Buyers vs Sellers Margin", "Buying pressure minus selling pressure over the last 14 "
                                     "days - the scoreboard of the daily tug-of-war. HOW TO "
                                     "READ IT: +15 or more = buyers clearly in charge (SJM was "
                                     "+18 the day it was bought). Low single digits = buyers "
                                     "barely ahead - the trend looks strong but nobody really "
                                     "owns it (MLI's +4 tell). Near zero = stalemate, a yellow "
                                     "flag on a fresh buy. Negative on a buy signal = sellers "
                                     "still winning; the gates block these. CAVEAT, measured "
                                     "2026-07-28 on 170 signals: a big margin did NOT predict "
                                     "better returns (slightly backwards). Read it like a fuel "
                                     "gauge, not a map: it describes the current state of who "
                                     "controls the stock - never a reason by itself to buy, "
                                     "rank higher, or pay up."),
        ("Trend Strength (ADX)", "Technically a 0-100 scale but real readings almost never "
                                 "get near 100, so do NOT read it as a percent. HOW TO READ "
                                 "IT: below 20 = no trend, sideways chop (our gates require "
                                 "20+). 20-25 = a trend is forming. 25-40 = a genuine, "
                                 "tradeable trend - most good entries live here; 35 is "
                                 "STRONG, not mediocre. 40-60 = powerful and often late. "
                                 "60+ = rare (panics and parabolas). It measures strength "
                                 "only, not direction - a crash reads high too. And it is "
                                 "SLOW (built from smoothed 14-day averages): it confirms a "
                                 "trend that exists, it does not predict one. A gate and a "
                                 "ranking aid, never a buy signal by itself."),
        ("Trend Strength vs Today's Best", "Each name's ADX as a share of the day's strongest "
                                           "(100 = strongest today). Fixes the read-as-percent "
                                           "trap above: a 35 that is today's best reads 100. "
                                           "Trend strength is the one measured-to-matter "
                                           "input, so this ranks the batch by it at a glance."),
        ("Magical (CCI-20)", "Above +100 = overbought, below -100 = oversold. NOTE: measured "
                             "2026-07-22 across 1,725 signals, this cut does NOT separate "
                             "returns (0.03pt over 21 days). Treat as context, not a veto."),
        ("ZLSMA-50", "The trend line. Price below it on a buy signal is a structural fail."),
        ("Chandelier Exit", "Flips BUY/SELL. The label price is the STOP level, not an entry."),
        ("Regime PASS", "Above a rising 200-day and beating SPY -> normal size."),
        ("Regime REPAIR", "Below the 200-day -> starter size only."),
        ("Regime DEEP-FAIL", "More than 10% below the 200-day -> watch only, no size."),
    ],
)

# ------------------------------------------------------- 1 Market & Rotation ---
ws = wb.create_sheet("1 - Market & Rotation")
style_header_row(ws, ["Indicator", "Reading", "What This Means For Today"],
                 ["The market measure being quoted.",
                  "Its current value.",
                  "Plain-English read: what this changes about how you trade today."],
                 [34, 18, 92], freeze="A2", autofilter=False)
for row in ctx.get("market_lines", []):
    ws.append([safe(x) for x in row])

# Derive the market read from the scan itself + the radar/ignition state files,
# rather than leaving the tab empty waiting for a hand-written context file.
_ok_all = res.get("all_names") or []
_sell = len(res.get("sell_mode") or [])
_tot = len(_ok_all) or res.get("scanned") or 0
if _tot:
    _pct_sell = 100.0 * _sell / _tot
    ws.append(["Breadth - names in sell mode", f"{_sell}/{_tot} ({_pct_sell:.0f}%)",
               "Over half the market in downtrend means long setups are swimming "
               "against the tide - size down or wait."
               if _pct_sell >= 50 else
               "Most names are not in downtrend - a normal tape for long setups."])
_fresh_n = res.get("fresh_count") or 0
_pass_n = sum(1 for h in res.get("hits", [])
              if str(h.get("verdict") or "").startswith(("CANDIDATE", "STARTER")))
ws.append(["Fresh buy signals", str(_fresh_n),
           "Raw Chandelier flips before any gate."])
ws.append(["Signals surviving all gates", str(_pass_n),
           "What is actually tradeable. If this is 0, the honest answer is no trades."])
if _fresh_n:
    ws.append(["Signal survival rate", f"{100.0*_pass_n/_fresh_n:.0f}%",
               "Low survival means the indicator is firing into a hostile tape."])

# --- Market structure (Omar 2026-07-28): the day's PATH, the close-within-range
# tell, rebound-vs-giveback and leadership concentration. A sector list with no
# daily change and no sense of HOW the day traded is data, not a read.
# Produced by scripts/market_structure.py; absent = tab degrades, never blocks.
try:
    MS = json.load(open(os.path.join(
        REPO, "watchlists", f"market-structure-{DATE}.json"), encoding="utf-8"))
except Exception:
    MS = None

if MS:
    _p = MS.get("index_path") or {}
    ws.append([])
    ws.append(["HOW THE DAY ACTUALLY TRADED"])
    ws.cell(row=ws.max_row, column=1).font = SECTION
    if _p.get("shape"):
        ws.append(["The session's path", safe(_p["shape"]),
                   "The closing number alone hides the shape. A flat close after a "
                   "big intraday swing is a very different day from a quiet one."])
    if _p.get("close_quality") is not None:
        ws.append(["Where it closed in its range",
                   f"{_p.get('range_pos')}% of the day's range",
                   safe(_p["close_quality"]) + " - close near the low means sellers "
                   "never let up, near the high means buyers pressed."])
    if _p.get("range_pct") is not None:
        ws.append(["Day's full range", f"{_p['range_pct']:.2f}%",
                   f"High {_p.get('high_pct'):+.2f}% / low {_p.get('low_pct'):+.2f}% "
                   "against the prior close."])
    _c = MS.get("concentration") or {}
    if _c:
        ws.append(["Groups up / down today",
                   f"{_c.get('n_up_today')} up, {_c.get('n_down_today')} down "
                   f"of {_c.get('groups')}",
                   "Breadth ACROSS GROUPS. A flat index with most groups up means "
                   "a few heavyweights did the damage, not broad selling."])
        ws.append(["Leadership concentration",
                   f"{_c.get('igniting_count')} group(s) clearly ahead "
                   f"(spread {_c.get('accel_spread')})",
                   safe(_c.get("read", ""))])
    _u = MS.get("unwind") or {}
    if _u.get("verdict"):
        ws.append(["Was this rotation or unwinding?", _u["verdict"].split(" - ")[0],
                   safe(_u["verdict"])])
        ws.append(["  - day's 3 best",
                   safe(", ".join(_u.get("top3", []))),
                   safe("Character: " + ", ".join(_u.get("top3_character", [])))])
        ws.append(["  - day's 3 worst",
                   safe(", ".join(_u.get("bottom3", []))),
                   safe("Character: " + ", ".join(_u.get("bottom3_character", [])))])

    # The judgement slot. Deliberately NOT auto-generated - a script naming what a
    # day "meant" would be confident noise. Written by hand with the EOD brief.
    ws.append([])
    ws.append(["NARRATIVE - written by hand each evening"])
    ws.cell(row=ws.max_row, column=1).font = SECTION
    _narr_path = os.path.join(REPO, "watchlists", f"market-narrative-{DATE}.md")
    if os.path.exists(_narr_path):
        for _ln in open(_narr_path, encoding="utf-8").read().splitlines():
            if _ln.strip():
                ws.append(["", safe(_ln.strip())])
    else:
        ws.append(["", "NOT WRITTEN YET - see the EOD brief for the written read. "
                       f"(Drop it in watchlists/market-narrative-{DATE}.md to have "
                       "it appear here.)"])

    # Sector table sorted by TODAY's move, with the benchmark left in place so
    # out/under-performance is readable at a glance.
    ws.append([])
    ws.append(["SECTORS BY TODAY'S MOVE"])
    ws.cell(row=ws.max_row, column=1).font = SECTION
    _hdr = ws.max_row + 1
    ws.append(["Group", "Today %", "Close in range %", "Volume vs normal",
               "1 Month %", "6 Month %", "Character"])
    for c in ws[_hdr]:
        c.font = BOLD
    _BENCH = MS.get("benchmark")
    for r in MS.get("rows", []):
        ws.append([safe(r.get("group")), r.get("day_pct"), r.get("range_pos"),
                   r.get("vol_x"),
                   round(r["m1"], 1) if r.get("m1") is not None else None,
                   round(r["m6"], 1) if r.get("m6") is not None else None,
                   safe(r.get("character"))])
        if r.get("ticker") == _BENCH:
            for c in ws[ws.max_row]:
                c.font = BOLD
    ws.append([])
    ws.append(["", safe("Character: REBOUND = up lately but still underwater over 6 "
                        "months (climbing out of a hole, not leading). GIVE-BACK = "
                        "down lately after a big 6-month run. LEADING = up on both. "
                        "LAGGING = neither. When the day's winners are REBOUND and "
                        "the losers are GIVE-BACK, money is unwinding what worked "
                        "rather than entering a new theme.")])
    ws.append(["", safe("Close in range: 0% = closed on the low (sellers had the last "
                        "word), 100% = closed on the high. Volume vs normal: 1.0 = an "
                        "average day; 2.0+ means the move had real conviction.")])

# Pull the headline lines straight out of the radar / ignition state files.
def _state_headlines(path, limit=8, section_kw="NEW"):
    """Pull the bullet lines out of a state file's 'NEW since prior run' section.

    First version stripped '-*' from the line start, which destroyed the '**'
    it then tested for, so it silently returned nothing and the tab read
    '(state not found)'. Match the section, then take bullets.
    """
    try:
        txt = open(os.path.join(REPO, path), encoding="utf-8").read()
    except OSError:
        return []
    lines, out, in_sec = txt.splitlines(), [], False
    for ln in lines:
        st = ln.strip()
        if st.startswith("##"):
            in_sec = section_kw.upper() in st.upper()
            continue
        if not in_sec or not st:
            continue
        if st.startswith(("-", "*", "·", "•")):
            out.append(re.sub(r"\*\*", "", st.lstrip("-*·• ").strip()))
        if len(out) >= limit:
            break
    return out

ws.append([])
ws.append(["IGNITION SWEEP - new since prior run"])
ws.cell(row=ws.max_row, column=1).font = SECTION
_ig = _state_headlines("watchlists/ignition-sweep-state.md", 12)
if _ig:
    for line in _ig:
        ws.append(["", safe(line)])
else:
    ws.append(["", "(ignition sweep state not found - see Data Quality tab)"])

ws.append([])
ws.append(["ROTATION RADAR"])
ws.cell(row=ws.max_row, column=1).font = BOLD
# Join the radar's ticker->state map onto market_structure's numbers (same universe,
# same source) when the rich shape is absent - which is always. See the parse note.
if not radar_lines and radar_states and MS:
    _by_tk = {r.get("ticker"): r for r in MS.get("rows", [])}
    for _tk, _state in radar_states.items():
        _r = _by_tk.get(_tk)
        if not _r:
            continue
        radar_lines.append({"group": _r.get("group", _tk), "state": _state,
                            "accel": _r.get("accel"),
                            "w": round(_r["w"], 1) if _r.get("w") is not None else None,
                            "m1": round(_r["m1"], 1) if _r.get("m1") is not None else None,
                            "m6": round(_r["m6"], 1) if _r.get("m6") is not None else None})
    radar_lines.sort(key=lambda x: -(x["accel"] if x["accel"] is not None else -999))
if radar_lines:
    hdr = ws.max_row + 1
    ws.append(["Group", "State", "Acceleration", "1 Week %", "1 Month %", "6 Month %"])
    for c in ws[hdr]:
        c.font = BOLD
    for r in radar_lines:
        ws.append([safe(r["group"]), safe(r["state"]), r["accel"], r["w"], r["m1"], r["m6"]])
        fill = GREEN if r["state"] == "IGNITING" else AMBER if r["state"] == "ROLLING" else None
        if fill:
            for c in ws[ws.max_row]:
                c.fill = fill
else:
    ws.append(["(rotation radar did not run - see Data Quality tab)"])

# The radar was MEASURED and came out inverted - groups it flagged as heating up
# did not go on to outperform. It must never be presented as a buy signal, and the
# caveat travels WITH the table so it cannot be read without it. (Omar 2026-07-28)
ws.append([])
ws.append(["", safe("READ THIS BEFORE ACTING ON THE TABLE ABOVE: these states were "
                    "back-tested and came out BACKWARDS - groups flagged as heating "
                    "up did NOT go on to outperform. Two same-week proofs: energy "
                    "went IGNITING on 7/24 and was demoted one session later on "
                    "7/27; airlines were demoted to ROLLING on 7/27, the day "
                    "airlines were the second-best group in the market at +2.9%. "
                    "Use this to see WHERE YOU ARE EXPOSED. Never as a reason to "
                    "buy or sell.")])
for row in ctx.get("rotation_commentary", []):
    ws.append([safe(row[0] if isinstance(row, list) else row)])

# ------------------------------------------------------------ 2 Fresh Buys ---
# ---- Catalyst calendar join (Omar 2026-07-23): catalyst as columns on the ticker rows,
# not a standalone tab. Fed by research/catalyst-calendar.json (build_catalyst_calendar.py). ----
try:
    _cat_raw = json.load(open(os.path.join(REPO, "research", "catalyst-calendar.json")))
    CATALYST = {r["ticker"].upper(): r for r in _cat_raw.get("records", [])}
except Exception:
    CATALYST = {}

CAT_COL = Col("catalyst", "Catalyst", 22,
              "Next scheduled catalyst for this name - earnings within 7 days, or a Phase-3 "
              "trial readout window. Blank = none flagged. From build_catalyst_calendar.py.")
CATVIEW_COL = Col("catview", "Catalyst View", 62,
                  "What the catalyst is and the suggested ACTION. Earnings: trim into the print, "
                  "re-enter post-print only if it beats AND holds (PEAD). Biopharma: BINARY - "
                  "defined-risk / awareness only, never hold a momentum long through it.")


def _inject_cat(rows):
    """Attach Catalyst + Catalyst View onto each ticker row before rendering."""
    for r in rows:
        sym = (r.get("sym") or r.get("a") or r.get("ticker") or "").upper()
        c = CATALYST.get(sym)
        r["catalyst"] = f"{c['event_type']} {c['date_or_window']}" if c else ""
        r["catview"] = c["view"] if c else ""


FRESH_COLS = [
    Col("rank", "RANK", 7,
        "Priority order, 1 = look at this first. Ranked by the composite BUY SCORE. "
        "UNCALIBRATED - it is a transparent weighting of things that ought to matter, "
        "NOT a measured edge. Use it to decide what to examine first, never as the "
        "reason to take a trade."),
    Col("buy_score", "BUY SCORE v2", 12,
        "0-100 rank, re-weighted 2026-07-22 after calibration: trend (ADX/DI) 45, "
        "structure (vs ZLSMA) 35, regime 15, freshness/liquidity 5. Risk quality "
        "was REMOVED from the sort - it correlated NEGATIVELY with returns (-0.05) "
        "and now shows only in the Risk column. STILL UNCALIBRATED: even v2 only "
        "rank-correlates ~+0.1-0.2 with forward returns and did NOT beat v1 out of "
        "sample. Use it to decide what to look at first, never as a forecast.", FMT_PCT),
    Col("risk_flag", "Risk (stop x ATR)", 15,
        "Stop distance in ATR multiples - the sizing/survival read, kept OUT of the "
        "ranking because it does not predict returns. 1.0-4.0x is healthy; below 1x "
        "gets noise-stopped, above 4x is a token position."),
    Col("sym", "Ticker", 9, "The stock symbol."),
    Col("company", "Company", 30, "Company name, so you know what you are buying."),
    # Omar 2026-07-28, after the HAS-vs-MLI post-mortem: the buyer/seller margin
    # and relative trend strength were the Friday-visible tells that ranked the
    # week's winners - and neither was a visible column. Now they are.
    Col("di_margin", "Buyers vs Sellers Margin", 13,
        "Buying pressure minus selling pressure (+DI minus -DI, 14-day) - who "
        "is winning the daily fight. In the week of 7/24 the wide-margin names "
        "(SJM +18) outran the thin-margin one (MLI +4). BUT MEASURED 2026-07-28 "
        "on 170 independent signals: the margin did NOT predict forward returns "
        "- slightly BACKWARDS (-0.12), while raw trend strength stayed the only "
        "positive input. So this is CONTEXT ONLY: it never ranks, never vetoes. "
        "Full test: research/di-margin-test-2026-07-28.md.", FMT_PCT),
    Col("adx_rel", "Trend Strength vs Today's Best", 13,
        "This name's trend strength (ADX) as a share of the STRONGEST trend "
        "among today's fresh signals - 100 means it is today's strongest. "
        "Trend strength is the one input our calibration found with real (if "
        "weak) predictive signal, so this ranks the batch by the thing "
        "measured to matter. On 2026-07-24 it would have read: HAS 100, MLI "
        "84, SJM 71, DVN 61 - the week then played out in nearly that order.", FMT_INT),
    Col("sector", "Sector", 20, "Sector - use it to spot when several signals are the same bet."),
    Col("signal_date", "Signal Date", 12, "The session the Chandelier flipped to BUY."),
    Col("age", "Sessions Old", 11,
        "How many sessions ago the signal fired. 0 = today's bar, which can still "
        "repaint before the close. Anything over 5 is not fresh."),
    Col("last", "Price", 10, "REAL last traded price from the live feed - what you can actually deal at. NOT the Heikin-Ashi bar value, which is a smoothed average and ran a median 0.59% (up to 5.3%) away from the true close.", FMT_PRICE),
    Col("stop", "Chandelier Stop", 13,
        "The Chandelier Exit level. This is the STOP, not an entry price - a common "
        "misreading. Price minus this level is your risk per share.", FMT_PRICE),
    Col("risk_pct", "Stop Distance %", 13,
        "How far the stop sits below price, as a percent. This decides your position "
        "size. Anything much over ~10% means a tiny position or no trade.", FMT_PCT),
    Col("magical", "Overbought / Oversold (CCI-20)", 16,
        "20-period CCI. Above +100 = stretched, below -100 = washed out. Measured "
        "2026-07-22: this does NOT predict returns, so it is context and never the "
        "sole reason to reject a name.", FMT_PCT),
    Col("zlsma", "Trend Line (ZLSMA-50)", 14,
        "Zero-lag moving average. Price BELOW this on a buy signal is a structural "
        "failure - the trend has not turned yet.", FMT_PRICE),
    Col("rsi", "Heat Gauge (RSI)", 11, "Daily RSI. Over 70 is hot, under 30 is cold.", FMT_PCT),
    Col("relvol", "Volume vs Normal", 12,
        "Today's volume against its average. 2.0 = twice normal interest. Real moves "
        "come with volume; a breakout on dry volume usually fails.", FMT_X),
    Col("pct52", "% of 52-Week High", 13,
        "100% means at the yearly high. Leaders live within about 15% of it.", FMT_PCT),
    Col("pct200", "% vs 200-Day Average", 14,
        "Above zero = long-term uptrend. Below = you are buying inside a downtrend.", FMT_PCT),
    Col("adx", "Trend Strength (ADX)", 12,
        "Above 20-25 means a real trend exists. Below that the move is noise.", FMT_PCT),
    Col("score", "Origination Score", 12,
        "The origination scanner's 0-100 quality score, if it also found this name.", FMT_PCT),
    Col("grade", "Origination Grade", 12, "A+ / A / B / C / SKIP from the origination scan."),
    Col("regime", "Regime", 12, "PASS = normal size, REPAIR = starter only, DEEP-FAIL = no size."),
    Col("verdict", "VERDICT", 26,
        "The decision the frozen rule set reached. Green = actionable, amber = "
        "conflicted (half size at most), red = blocked."),
    Col("why", "Why (short)", 55,
        "The one-line reason for the verdict. The FULL argument is on the "
        "'7 - Notes & Decisions' tab - this column is deliberately short so the "
        "table stays readable."),
]


def _tv_screen_metrics():
    """RSI / volume-vs-normal / 52-week-high distance for the whole US market,
    one public-scanner call, mapped by bare ticker.

    Omar 2026-07-28: the Heat Gauge (RSI), Volume vs Normal and % of 52-Week
    High columns sat EMPTY on every chart-read run - they were only ever filled
    from the origination enrich file, which chart-read runs do not produce.
    The public scanner (already used by rotation_radar / live_picks_radar)
    has all three. Non-fatal: no network means the columns stay blank and a
    Data Quality row says so, instead of the blanks passing as normal.
    """
    import urllib.request
    body = json.dumps({
        "filter": [{"left": "type", "operation": "equal", "right": "stock"},
                   {"left": "exchange", "operation": "in_range",
                    "right": ["NASDAQ", "NYSE", "AMEX"]}],
        "options": {"lang": "en"}, "markets": ["america"],
        "columns": ["name", "RSI", "relative_volume_10d_calc",
                    "close", "price_52_week_high"],
        "range": [0, 9000],
    }).encode()
    req = urllib.request.Request(
        "https://scanner.tradingview.com/america/scan", data=body,
        headers={"User-Agent": "Mozilla/5.0", "Content-Type": "application/json"})
    out = {}
    for r in json.loads(urllib.request.urlopen(req, timeout=45).read()).get("data", []):
        d = r.get("d") or []
        if len(d) < 5:
            continue
        bare = str(r.get("s", "")).split(":")[-1]
        rsi, rvol, close, hi52 = d[1], d[2], d[3], d[4]
        out[bare] = {
            "rsi": round(rsi, 1) if isinstance(rsi, (int, float)) else None,
            "relvol": round(rvol, 2) if isinstance(rvol, (int, float)) else None,
            "close": round(close, 2) if isinstance(close, (int, float)) else None,
            "pct52": round(close / hi52 * 100, 1)
                     if (isinstance(close, (int, float)) and isinstance(hi52, (int, float)) and hi52) else None,
        }
    return out


try:
    TV_METRICS = _tv_screen_metrics()
    _TV_METRICS_NOTE = None
except Exception as _exc:
    TV_METRICS = {}
    _TV_METRICS_NOTE = str(_exc)


SCORED = {}   # sym -> (score, components, pros, cons); reused by Notes & Plans


def fresh_rows():
    out = []
    for sym, h in hits.items():
        e, o = enrich.get(sym, {}), orig.get(sym, {})
        last, stop = h.get("last"), h.get("ce_label") or h.get("flip_level")
        # SIGNAL side keeps the Heikin-Ashi numbers above (every gate and score is
        # computed on them and that maths is internally consistent).
        # DISPLAY side uses the REAL, tradeable prices - Omar's 2026-07-25 call.
        # `last` is an HA close, a smoothed (O+H+L+C)/4 average that nobody can
        # trade at; it ran a median 0.59% away from the true close and up to 5.3%.
        # Quoting it as an entry put that error straight into the plan.
        d_last, d_stop, d_risk = plan_display(h, last, stop)
        _, why, _ = split_note(h.get("note"))
        why = plainify(why)

        g = h.get("gates") or {}
        feat = {
            "last": last, "long_stop": stop, "zlsma": h.get("zlsma"),
            "atr": h.get("atr") or g.get("atr"), "adx": g.get("adx"), "plus_di": g.get("plus_di"),
            "minus_di": g.get("minus_di"), "regime": g.get("regime") or h.get("regime"),
            "pct_vs_200": g.get("pct_vs_200"), "avg_dollar_vol": g.get("avg_dollar_vol"),
            "bars_back": h.get("bars_back"),
        }
        # v2 ranks; v1 kept only for its pros/cons narrative (component breakdown).
        sc, _v2comp, rflag = score_long_v2(feat, position=DEFAULT_POSITION)
        _v1sc, comp, pros, cons = score_long(feat, position=DEFAULT_POSITION)
        # A blocked name can never outrank a live candidate, whatever it scores.
        if str(h.get("verdict") or "").startswith("BLOCKED"):
            sc = round(sc * 0.4, 1)
        SCORED[sym] = (sc, comp, pros, cons)
        risk_txt = (f"{rflag['stop_x_atr']}x ({rflag['stop_pct']}%)"
                    + ("" if rflag.get("ok") else " !") if rflag else "-")

        # Sweep gate data first, origination enrich second, public scanner third
        # - so these columns are populated on chart-read runs too (Omar 7/28).
        tv = TV_METRICS.get(sym, {})
        pdi, ndi = g.get("plus_di"), g.get("minus_di")
        adx_val = g.get("adx") if g.get("adx") is not None else e.get("adx")
        out.append({
            "buy_score": sc, "risk_flag": risk_txt,
            "sym": sym, "company": e.get("description"), "sector": e.get("sector"),
            "di_margin": (round(pdi - ndi, 1)
                          if pdi is not None and ndi is not None else None),
            "signal_date": h.get("signal_date_est"), "age": h.get("bars_back"),
            "last": d_last, "stop": d_stop, "risk_pct": d_risk,
            "magical": h.get("magical"), "zlsma": h.get("zlsma"),
            "rsi": (round(e["rsi"], 1) if e.get("rsi") else tv.get("rsi")),
            "relvol": (round(e["relvol"], 2) if e.get("relvol") else tv.get("relvol")),
            "pct52": e.get("pct_of_52wk_high") or tv.get("pct52"),
            "pct200": (g.get("pct_vs_200") if g.get("pct_vs_200") is not None
                       else e.get("pct_vs_200d")),
            "adx": round(adx_val, 1) if adx_val is not None else None,
            "score": o.get("score"), "grade": o.get("grade"),
            "regime": h.get("regime"), "verdict": h.get("verdict"), "why": why,
        })
    out.sort(key=lambda r: -(r["buy_score"] or 0))
    _max_adx = max((r["adx"] for r in out if r["adx"]), default=None)
    for i, r in enumerate(out, 1):
        r["rank"] = i
        r["adx_rel"] = (round(100 * r["adx"] / _max_adx)
                        if r["adx"] and _max_adx else None)
    return out


_fresh = fresh_rows()
ws = wb.create_sheet("2 - Fresh Buys")
_inject_cat(_fresh)
write_table(ws, FRESH_COLS + [CAT_COL, CATVIEW_COL], _fresh, row_fill=verdict_fill)

# ------------------------------------------------- 2b Gated Longs (all ages) ---
# Every BUY-mode name that clears the FULL gate stack, regardless of how long ago it
# flipped (Fresh Buys is <=5 bars only). Surfaces established uptrends like DVN/PBR
# that are still valid gated longs (Omar 2026-07-24). Ranked least-extended first.
GATED_COLS = [
    Col("rank", "RANK", 6, "Least-extended-above-the-200-day first - names with the most room "
        "and best entry reward-to-risk are at the top."),
    Col("sym", "Ticker", 9, "The stock symbol."),
    Col("last", "Price", 10, "REAL last traded price from the live feed - what you can actually deal at. NOT the Heikin-Ashi bar value, which is a smoothed average and ran a median 0.59% (up to 5.3%) away from the true close.", FMT_PRICE),
    Col("stop", "Chandelier Stop", 13, "The Chandelier trend-break level - the STOP, not an entry "
        "price. Price minus this is your risk per share.", FMT_PRICE),
    Col("stop_pct", "Stop %", 8, "How far the stop sits below price.", FMT_PCT),
    Col("ext", "Extension vs 200d", 15, "How far above the 200-day average. HIGH = stretched = "
        "worse entry. This is the extension label - the Fresh tab had no such view, which is how "
        "DVN/PBR stayed hidden.", FMT_PCT),
    Col("magical", "Overbought (CCI-20)", 14, "Above +100 = stretched. Measured NOT to predict "
        "returns - context only, never the sole reason to skip.", FMT_PCT),
    Col("adx", "Trend (ADX)", 11, "20+ = a real trend is present.", FMT_PCT),
    Col("age", "Sessions Old", 11, "Sessions since the Chandelier flipped BUY. Fresh Buys shows "
        "<=5; THIS tab shows every age."),
    Col("regime", "Regime", 10, "PASS = above a rising 200-day. REPAIR = below it, starter size only."),
    Col("note", "Gate Notes", 44, "Why it passed / any starter flag."),
]
_gated = res.get("gated_longs", [])
grows = []
for _i, _g in enumerate(_gated, 1):
    _last, _stop, _spct = plan_display(_g, _g.get("last"), _g.get("stop"))
    grows.append({
        "rank": _i, "sym": _g["sym"], "last": _last, "stop": _stop,
        "stop_pct": _spct,
        "ext": round(_g["pct_vs_200"], 1) if _g.get("pct_vs_200") is not None else None,
        "magical": _g.get("magical"), "adx": _g.get("adx"), "age": _g.get("bars_back"),
        "regime": _g.get("regime"),
        "note": plainify(("STARTER (regime REPAIR - half size). "
                          if str(_g.get("verdict", "")).startswith("STARTER") else "")
                         + (_g.get("note") or "")),
    })
_inject_cat(grows)
ws = wb.create_sheet("2b - Gated Longs")
write_table(ws, GATED_COLS + [CAT_COL, CATVIEW_COL], grows)
ws.append([])
ws.append([safe(f"GATED LONGS (all ages): {len(grows)} names currently in BUY mode that clear the "
                "FULL gate stack - not just fresh flips (Fresh Buys is <=5 sessions). Ranked "
                "LEAST-EXTENDED first = best entry reward-to-risk on top.")])
ws.append([safe("This is where established-but-still-valid uptrends live (DVN, PBR and the like). "
                "Check Extension + Overbought + the Catalyst columns before acting - many "
                "established uptrends are stretched or reporting earnings within days.")])
ws.append([])
ws.append([safe("RANKED STRONGEST -> WEAKEST by BUY SCORE. The score is UNCALIBRATED - "
                "a transparent weighting, not a measured edge. Blocked names are scored "
                "down so they can never outrank a live candidate.")])
ws.cell(row=ws.max_row, column=1).font = BOLD
for k, v in RANK_RATIONALE.items():
    ws.append([safe(f"  {k}"), safe(v)])

# ---------------------------------------------------------------- 3 Plans ---
# Column widths + the medium-border grid below are Omar's hand-layout of the
# Plans tab in universe_2026-07-27_new.xlsx (2026-07-28): number columns tight,
# Thesis & Break the dominant reading column, every cell boxed.
PLAN_COLS = [
    Col("a", "Ticker", 6.1, "The stock symbol."),
    Col("b", "Type", 9.9, "Order type - buy-stop, limit, or starter."),
    Col("c", "Entry", 10.1, "The price you are willing to pay.", FMT_PRICE),
    Col("d", "Stop", 9.6, "Where the trade is wrong and you are out.", FMT_PRICE),
    Col("e", "Size", 9.1, "Dollar size and share count."),
    Col("f", "Risk $", 10.6, "Dollars lost if the stop hits. Should be about 0.5% of the account.", FMT_PRICE),
    Col("g", "Checkpoint", 10.9,
        "NOT a sell level. The old 2:1 target, kept as a progress checkpoint "
        "(retired 2026-07-25: targets measured harmful - they cap the tail winners). "
        "Exits are: stop / thesis break / 20% ratchet below the highest close once "
        "well in profit."),
    Col("h", "Earnings Gate", 28.7,
        "Next earnings date and whether it blocks or shrinks this trade. Never full "
        "size into a report. Fed by the catalyst calendar; 'Clear' still means "
        "confirm the date yourself."),
    Col("i", "Alert", 41.9,
        "The alert IDs armed for this plan, or the exact price levels to wire on "
        "entry (entry limit + trend-break stop) and the current board state."),
    Col("j", "Plan Notes", 32.6, "Conditions attached to the plan - what must be true to take it."),
    Col("k", "Thesis & Break", 85.7,
        "What this trade is betting on, and the observation that proves it wrong "
        "(protocol rule 11, added 2026-07-26). A thesis BREAK = exit regardless of "
        "stop or profit. A plan without a break condition is not takeable."),
]
# Build plans from the gated candidates. Previously this tab was fed only by a
# hand-written context file, so it was almost always EMPTY - which read as "no
# work done" rather than "no name qualified". Now it is always populated: either
# with real sized plans, or with an explicit statement of why there are none.
ACCOUNT, RISK_UNIT, MAX_ORDER = 1_000_000.0, 5_000.0, 100_000.0


def _earn_txt(sym, h):
    """Earnings cell - say what is actually KNOWN, not a bare 'verify'.

    Omar 2026-07-28: this column sat useless ('NOT CHECKED') on every row while
    the catalyst calendar - already built and already joined onto three other
    tabs - had the answer. Use it.
    """
    e = h.get("earnings") or {}
    if e:
        est = e.get("estimate", "?")
        tag = "CONFIRMED" if e.get("confirmed") else "ESTIMATED"
        days = e.get("days_out")
        out = f"{est} ({tag}"
        if days is not None:
            out += f", ~{days}d out"
        out += ")"
        if e.get("rule"):
            out += f" - {e['rule']}"
        return out
    c = CATALYST.get(sym)
    if c and str(c.get("event_type", "")).lower().startswith("earn"):
        when = c.get("date_or_window", "?")
        try:
            days = (datetime.date.fromisoformat(when)
                    - datetime.date.fromisoformat(DATE)).days
            when_txt = f"{when} ({days} day{'s' if days != 1 else ''} away)"
        except ValueError:
            when_txt = when
        detail = c.get("detail")
        return (f"REPORTS {when_txt}" + (f" - {detail}" if detail else "")
                + ". Inside the 7-day window: do NOT carry full size into the "
                  "print. Fresh entry only after it reports, beats, and holds.")
    if c:
        return (f"{c.get('event_type')} {c.get('date_or_window')} - a binary "
                "event; see the Catalyst View column on the Fresh Buys tab.")
    return ("Clear - no earnings on the 7-day catalyst calendar. Still confirm "
            "the date yourself before entry; the calendar has missed changes.")


def _alert_txt(sym, h, entry=None, stop=None):
    """Alert cell - the LEVELS that matter for this stock, and the board state.

    Omar 2026-07-28: a bare 'NOT ARMED' says nothing. Show the two prices an
    alert should sit at (entry limit + trend-break stop), and whether the board
    already has something active on the name so old alerts get retired first.
    """
    a = h.get("alerts") or {}
    if a:
        return "; ".join(f"{k} {v.get('level')} (id {v.get('id')})"
                         for k, v in a.items())
    b = ALERT_BOARD.get(sym)
    board = ("the board already has an ACTIVE alert on this name - re-read its "
             "message before wiring more, retire it if stale"
             if b and b.get("active") else "nothing on the board yet")
    levels = []
    if entry:
        levels.append(f"entry limit at {entry:.2f}")
    if stop:
        levels.append(f"trend-break at {stop:.2f}")
    lvl = " and ".join(levels) if levels else "levels TBD"
    return (f"WIRE ON ENTRY: {lvl}. Right now {board}. Every alert message "
            "must say what the level means and what to do when it fires.")


try:
    ALERT_BOARD = {r["sym"]: r for r in json.load(
        open(os.path.join(REPO, "watchlists", "alert-board.json"), encoding="utf-8"))}
except Exception:
    ALERT_BOARD = {}


def _thesis_txt(r):
    """Thesis & Break cell, written the way Omar reads: what the bet is, what
    proves it wrong, and what still has to be added by hand before entry."""
    g = hits.get(r["sym"], {}).get("gates") or {}
    regime = g.get("regime") or r.get("regime") or "?"
    adx = g.get("adx") if g.get("adx") is not None else r.get("adx")
    adx_txt = f"about {adx:.0f}" if isinstance(adx, (int, float)) else "unknown"
    return (f"The bet: this stock's uptrend just resumed and should keep going. "
            f"It is above its 200-day average (health check: {regime}), the "
            f"trend-strength reading is {adx_txt}, and the Chandelier just "
            f"flipped to BUY. The bet is WRONG the day the daily Chandelier "
            f"flips back to SELL, or the health check stops passing - exit that "
            f"day, even if the stop has not been hit and even if the trade is "
            f"in profit. Before entering, add the one company-specific reason "
            f"this should work (the catalyst, or the level that must hold). "
            f"A plan without that is not takeable - protocol rule 11.")


plan_rows = []
for r in _fresh:
    v = str(r.get("verdict") or "")
    if not v.startswith(("CANDIDATE", "STARTER")):
        continue
    last, stop = r.get("last"), r.get("stop")
    if not (last and stop and last > stop):
        continue
    risk_ps = last - stop
    shares = int(min(MAX_ORDER / last, RISK_UNIT / risk_ps))
    dollars = shares * last
    starter = v.startswith("STARTER")
    if starter:                      # regime REPAIR -> half size
        shares, dollars = shares // 2, (shares // 2) * last
    target = last + 2 * risk_ps
    plan_rows.append({
        "a": r["sym"],
        "b": "LIMIT (starter)" if starter else "LIMIT",
        "c": round(last, 2), "d": round(stop, 2),
        "e": f"${dollars:,.0f} ({shares:,} sh)",
        "f": round(shares * risk_ps, 0),
        "g": f"{target:.2f} CHECKPOINT (no sell)",
        "h": _earn_txt(r["sym"], hits.get(r["sym"], {})),
        "i": _alert_txt(r["sym"], hits.get(r["sym"], {}), last, stop),
        "j": (f"Rank #{r['rank']} of {len(_fresh)}, buy score {r['buy_score']}. "
              f"Stop is the Chandelier level ({r.get('risk_pct')}% away). "
              + ("HALF SIZE: regime REPAIR. " if starter else "")
              + "Check the stop against a normal day's swing by hand - the ATR "
                "floor is not wired into the gate stack yet."),
        "k": _thesis_txt(r),
    })

ws = wb.create_sheet("3 - Plans")
write_table(ws, PLAN_COLS, plan_rows)
# Omar's grid: a medium border on every cell of the table, header included.
grid_borders(ws, 1, 1 + len(plan_rows), 1, len(PLAN_COLS))
if not plan_rows:
    ws.append([])
    ws.append([safe("NO PLANS - and that is a result, not an omission.")])
    ws.cell(row=ws.max_row, column=1).font = RED_FONT
    ws.append([safe(f"{len(_fresh)} fresh buy signals were scanned. None survived the "
                    f"gate stack (regime, ADX, DI direction, ZLSMA, liquidity), so "
                    f"there is nothing to size.")])
    ws.append([safe("Do NOT relax a gate to make this tab non-empty. See the Blocked "
                    "tab for exactly which gate stopped each name.")])
ws.append([])
ws.append([safe(f"Sizing basis: ${ACCOUNT:,.0f} account, ${RISK_UNIT:,.0f} risk per "
                f"trade (0.5%), ${MAX_ORDER:,.0f} max order. Shares = the SMALLER of "
                f"the order cap and the risk unit divided by stop distance.")])

# -------------------------------------------------------------- 4 Blocked ---
BLOCK_COLS = [
    Col("rank", "RANK", 7,
        "1 = WORST, i.e. furthest from ever being tradeable. The top of this list is "
        "what to stop looking at; the BOTTOM is what is nearly there and worth "
        "re-checking tomorrow."),
    Col("severity", "How Badly It Fails", 13,
        "Severity score. Higher = more gates failed and failed harder. Built from: "
        "regime DEEP-FAIL 40 / REPAIR 15, no trend 20, sellers in control 15, "
        "below ZLSMA 15, illiquid 25, plus distance below the 200-day.", FMT_PCT),
    Col("sym", "Ticker", 9, "The stock symbol."),
    Col("company", "Company", 30, "Company name."),
    Col("signal_date", "Signal Date", 12, "When the buy signal fired."),
    Col("magical", "Overbought / Oversold (CCI-20)", 16, "20-period CCI at the read.", FMT_PCT),
    Col("regime", "Regime", 12, "PASS / REPAIR / DEEP-FAIL."),
    Col("gate", "Which Gate Blocked It", 30,
        "The specific rule that rejected this name. If you disagree with the rule, "
        "this is the column that tells you which rule to argue with."),
    Col("why", "Why (short)", 60, "One-line explanation. Full reasoning on the Notes tab."),
]
block_rows = []
for sym, h in hits.items():
    v = str(h.get("verdict") or "")
    if v.startswith("BLOCKED"):
        e = enrich.get(sym, {})
        _, why, _ = split_note(h.get("note"))
        g = h.get("gates") or {}
        sev, _reasons = rank_blocked_severity({
            "regime": g.get("regime") or h.get("regime"), "adx": g.get("adx"),
            "plus_di": g.get("plus_di"), "minus_di": g.get("minus_di"),
            "last": h.get("last"), "zlsma": h.get("zlsma"),
            "avg_dollar_vol": g.get("avg_dollar_vol"), "pct_vs_200": g.get("pct_vs_200"),
        })
        block_rows.append({"severity": sev,
                           "sym": sym, "company": e.get("description"),
                           "signal_date": h.get("signal_date_est"),
                           "magical": h.get("magical"), "regime": h.get("regime"),
                           "gate": plainify(v.replace("BLOCKED: ", "")),
                           "why": plainify(why)})
block_rows.sort(key=lambda r: -(r["severity"] or 0))
for i, r in enumerate(block_rows, 1):
    r["rank"] = i

ws = wb.create_sheet("4 - Blocked")
_inject_cat(block_rows)
write_table(ws, BLOCK_COLS + [CAT_COL, CATVIEW_COL], block_rows,
            row_fill=lambda r: RED if (r["severity"] or 0) >= 60 else
                               (AMBER if (r["severity"] or 0) >= 30 else GREY))
ws.append([])
ws.append([safe("RANKED WORST FIRST. Rank 1 is the most comprehensively unsuitable; "
                "the names at the BOTTOM failed on one thing and are worth re-checking "
                "tomorrow.")])
ws.cell(row=ws.max_row, column=1).font = BOLD
ws.append([])
ws.append([safe("SHOULD YOU SHORT THESE? NO.")])
ws.cell(row=ws.max_row, column=1).font = RED_FONT
ws.append([safe("A failed long is not a good short. These names failed a test asking "
                "'is this going up?' - answering no is not the same as 'this is going "
                "down'. Many are blocked for reasons that say nothing about direction "
                "at all: illiquid, earnings too close, stop too tight. See the Sell "
                "Mode tab for the short discussion.")])

# ------------------------------------------------------------ 5 Sell Mode ---
SELL_COLS = [
    Col("rank", "SHORT RANK", 11,
        "1 = the most technically shortable name here. THIS IS NOT A RECOMMENDATION "
        "TO SHORT - read the policy block below the table. It ranks which names are "
        "in the cleanest downtrends, for exit/hedge decisions and for paper-trading "
        "a short book before risking money."),
    Col("short_score", "SHORT SCORE", 12,
        "0-100 mirror of the buy score: downtrend depth 30, trend strength 25 "
        "(ADX + DI spread the other way), structure below ZLSMA 15, stop quality 15, "
        "squeeze safety 15. UNCALIBRATED and, unlike the long side, never tested "
        "against forward returns at all.", FMT_PCT),
    Col("sym", "Ticker", 9, "The stock symbol."),
    Col("company", "Company", 32, "Company name."),
    Col("sector", "Sector", 22, "Sector."),
    Col("last", "Price", 10, "Latest price.", FMT_PRICE),
    Col("pct200", "% vs 200-Day Average", 14, "How far below its long-term average it is.", FMT_PCT),
    Col("adx", "Trend Strength (ADX)", 12, "Above 20 = a real downtrend, not chop.", FMT_PCT),
    Col("di", "-DI vs +DI", 12, "Sellers vs buyers. Sellers need to be clearly ahead."),
    Col("magical", "Overbought / Oversold (CCI-20)", 16, "20-period CCI.", FMT_PCT),
    Col("risk", "Main Short Risk", 46,
        "The specific reason this short could hurt you - squeeze, thinness, or being "
        "already too far gone."),
]
sell_rows = []
for s in res.get("sell_mode", []):
    a = allrows.get(s, {}) or {}
    row = {"last": a.get("last"), "zlsma": a.get("zlsma"),
           "short_stop": a.get("short_stop") or a.get("ce_label"),
           "adx": a.get("adx"), "plus_di": a.get("plus_di"),
           "minus_di": a.get("minus_di"), "pct_vs_200": a.get("pct_vs_200"),
           "avg_dollar_vol": a.get("avg_dollar_vol")}
    sc, comp, pros, cons = score_short(row, position=DEFAULT_POSITION)
    sell_rows.append({
        "short_score": sc, "sym": s,
        "company": enrich.get(s, {}).get("description"),
        "sector": enrich.get(s, {}).get("sector"),
        "last": a.get("last"), "pct200": row["pct_vs_200"], "adx": row["adx"],
        "di": (f"{row['minus_di']:.0f} vs {row['plus_di']:.0f}"
               if row["minus_di"] is not None and row["plus_di"] is not None else None),
        "magical": a.get("magical"),
        "risk": "; ".join(cons[:2]) if cons else "no specific flag - still unvalidated",
    })
sell_rows.sort(key=lambda r: -(r["short_score"] or 0))
for i, r in enumerate(sell_rows, 1):
    r["rank"] = i

ws = wb.create_sheet("5 - Sell Mode")
_inject_cat(sell_rows)
write_table(ws, SELL_COLS + [CAT_COL, CATVIEW_COL], sell_rows,
            row_fill=lambda r: RED if (r["short_score"] or 0) >= 70 else None)
ws.append([])
for line in SHORT_POLICY.strip().split("\n"):
    ws.append([safe(line)])
    if line.strip().startswith("SHOULD YOU SHORT"):
        ws.cell(row=ws.max_row, column=1).font = RED_FONT

# ------------------------------------------------------- 6 Tracker Broken ---
TRACK_COLS = [
    Col("sym", "Ticker", 9, "The stock symbol."),
    Col("rec_date", "Recommended On", 13, "The date of the origination pick."),
    Col("grade", "Grade", 8, "The origination grade the pick carried (A+/A/B/C)."),
    Col("since", "% Since Recommended", 16, "Return since the call was made.", FMT_PCT),
    Col("rec", "Recommended At", 13, "The price when it was recommended.", FMT_PRICE),
    Col("last", "Price Now", 11, "Latest real price from today's scan.", FMT_PRICE),
    Col("magical", "Overbought / Oversold (CCI-20)", 16, "20-period CCI.", FMT_PCT),
    Col("reason", "Why It Broke", 70, "What changed - the trend condition that failed, "
        "and what the playbook says to do about it."),
]


def _derive_tracker_broken():
    """Build the broken-picks list the scan pipeline never populated.

    BUG FOUND 2026-07-28 (Omar: 'there has not been any data populating'):
    build_universe_results.py has ALWAYS written "tracker_exit_watch": {} - a
    hardcoded empty dict - so this tab compiled to a header row and nothing else
    on every single run. Nothing ever computed it.

    The inputs were sitting in the repo the whole time: the origination
    recommendation log (snapshotted daily into research/) says what was picked,
    when, and at what price; today's scan says which of those names the
    Chandelier now has in SELL mode. A pick is BROKEN when the indicator has
    flipped against it - the uptrend the pick was riding is over.
    """
    import csv
    import glob as _glob
    snaps = sorted(_glob.glob(os.path.join(REPO, "research",
                                           "recommendations_log-*.csv")))
    usable = [s for s in snaps
              if re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(s))
              and re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(s)).group(1) <= DATE]
    if not usable:
        usable = snaps
    if not usable:
        return {"broken": [], "breakage_stat":
                "No origination recommendation log found in research/ - "
                "run scripts/track_record.py first.",
                "coverage": "none", "gap_note": "no log, nothing to check"}
    snap = usable[-1]
    rows = list(csv.DictReader(open(snap, encoding="utf-8", errors="replace")))
    open_rows = [r for r in rows if str(r.get("status", "")).upper() == "OPEN"]
    # One row per ticker: the most recent open call (a ticker re-picked three
    # times would otherwise clutter the tab with three identical breaks).
    latest = {}
    n_calls = {}
    for r in sorted(open_rows, key=lambda x: x.get("date") or ""):
        latest[r["ticker"]] = r
        n_calls[r["ticker"]] = n_calls.get(r["ticker"], 0) + 1
    sell_set = set(res.get("sell_mode") or [])
    scanned = {a["sym"]: a for a in res.get("all_names") or []}
    in_scan = [t for t in latest if t in scanned]
    broken = []
    for t, r in latest.items():
        if t not in sell_set:
            continue
        a = scanned.get(t, {})
        now = a.get("real_last") or a.get("real_close") or a.get("last")
        try:
            rec_price = float(r.get("price"))
        except (TypeError, ValueError):
            rec_price = None
        since = (round((now / rec_price - 1) * 100, 1)
                 if (now and rec_price) else None)
        stand = ("The gain that was on the table is being given back - bank "
                 "what is left rather than round-trip it."
                 if (since or 0) > 0 else
                 "It never worked and the trend is now against it - cut it "
                 "rather than hope.")
        extra = (f" This name has been picked {n_calls[t]} times; this grades "
                 f"the latest call." if n_calls[t] > 1 else "")
        broken.append({
            "sym": t, "rec_date": r.get("date"), "grade": r.get("grade"),
            "since": since, "rec": rec_price, "last": now,
            "magical": a.get("magical"),
            "reason": (f"Picked {r.get('date')} at ${rec_price:,.2f} as a "
                       f"grade-{r.get('grade')} name. The Chandelier has since "
                       f"flipped it to SELL - the uptrend the pick was riding "
                       f"is over. {stand}{extra}"
                       if rec_price else
                       f"Picked {r.get('date')}. The Chandelier has since "
                       f"flipped it to SELL - the uptrend is over."),
        })
    broken.sort(key=lambda x: x.get("since") if x.get("since") is not None else 0)
    n_open, n_in, n_broken = len(latest), len(in_scan), len(broken)
    pct = round(100.0 * n_broken / n_in) if n_in else 0
    return {
        "broken": broken,
        "breakage_stat": (
            f"{n_broken} of the {n_in} open origination picks visible in "
            f"today's scan ({pct}%) have BROKEN trend - the indicator that "
            f"justified them has flipped to SELL. Log: {os.path.basename(snap)}."),
        "coverage": f"{n_in} of {n_open} open picks appear in today's scan universe",
        "gap_note": ("picks outside the scan universe cannot be checked here - "
                     "absence from this tab is not a clean bill of health for them"),
    }


tw = res.get("tracker_exit_watch") or {}
if not tw.get("broken"):
    tw = _derive_tracker_broken()
track_rows = [{"sym": b["sym"], "rec_date": b.get("rec_date"), "grade": b.get("grade"),
               "since": b.get("since_rec_pct", b.get("since")), "rec": b.get("rec"),
               "last": b.get("last"), "magical": b.get("magical"),
               "reason": plainify(b.get("reason"))}
              for b in sorted(tw.get("broken", []),
                              key=lambda x: x.get("since_rec_pct") or x.get("since") or 0)]
ws = wb.create_sheet("6 - Tracker Broken")
write_table(ws, TRACK_COLS, track_rows,
            row_fill=lambda r: RED if (r.get("since") or 0) < 0 else AMBER)
if not track_rows:
    ws.append([])
    ws.append([safe("No open origination pick currently sits in SELL mode - "
                    "nothing has broken as of this run.")])
ws.append([])
ws.append([safe(tw.get("breakage_stat", ""))])
ws.cell(row=ws.max_row, column=1).font = RED_FONT
ws.append([safe(f"Coverage: {tw.get('coverage', '?')} - {tw.get('gap_note', '')}")])

# ------------------------------------------------------ 7 Notes & Decisions ---
# "What Changed" was empty on every row (Omar 2026-07-28): it only ever showed a
# tag if the scan note happened to START with "REVERSAL."/"DOWNGRADED." etc.,
# which the pipeline never writes. Compute it properly: compare this run's
# verdict per name against the PREVIOUS run's results file.
_prev_hits, _PREV_DATE = {}, None
_prev_files = sorted(
    f for f in os.listdir(os.path.join(REPO, "watchlists"))
    if re.match(r"universe-results-\d{4}-\d{2}-\d{2}\.json$", f) and f[17:-5] < DATE)
if _prev_files:
    _PREV_DATE = _prev_files[-1][17:-5]
    try:
        _pj = json.load(open(os.path.join(REPO, "watchlists", _prev_files[-1])))
        _prev_hits = {h["sym"]: str(h.get("verdict") or "") for h in _pj.get("hits", [])}
    except Exception:
        _prev_hits, _PREV_DATE = {}, None


def _verdict_class(v):
    v = str(v or "")
    if v.startswith("CANDIDATE"):
        return "actionable"
    if v.startswith("STARTER"):
        return "starter-size only"
    if v.startswith("BLOCKED"):
        return "blocked"
    return "unknown"


def _what_changed(sym, verdict, note_tag):
    if note_tag:                      # explicit tag written by the scan wins
        return note_tag
    if not _PREV_DATE:
        return "(no earlier run on file to compare against)"
    prev = _prev_hits.get(sym)
    if prev is None:
        return (f"NEW - this name was not a fresh signal in the last run "
                f"({_PREV_DATE}).")
    now_c, prev_c = _verdict_class(verdict), _verdict_class(prev)
    if now_c == prev_c:
        return f"Same call as {_PREV_DATE} - still {now_c}."
    order = {"blocked": 0, "starter-size only": 1, "actionable": 2}
    word = ("UPGRADED" if order.get(now_c, 0) > order.get(prev_c, 0)
            else "DOWNGRADED")
    return (f"{word} since {_PREV_DATE}: was {prev_c} "
            f"('{prev}'), now {now_c}.")


_rank_of = {r["sym"]: r["rank"] for r in _fresh}
note_rows = []
for sym, h in hits.items():
    changed, _, full = split_note(h.get("note"))
    sc, comp, pros, cons = SCORED.get(sym, (None, {}, [], []))
    note_rows.append({
        "rank": _rank_of.get(sym),
        "sym": sym, "verdict": h.get("verdict"),
        "changed": _what_changed(sym, h.get("verdict"), changed),
        "pros": plainify(("* " + "\n* ".join(pros)) if pros
                         else "(nothing argues for it)"),
        "cons": plainify(("* " + "\n* ".join(cons)) if cons
                         else "(no flags - genuinely clean)"),
        "reasoning": plainify(full) or "(no narrative recorded)",
    })
note_rows.sort(key=lambda r: (r["rank"] is None, r["rank"] or 999))
notes_sheet(wb.create_sheet("7 - Notes & Decisions"), note_rows)

# ---------------------------------------------------------- 8 Market Movers ---
# The day's biggest gainers/losers across the ENTIRE market, each run through the
# protocol. Requested by Omar 2026-07-22. Data from watchlists/market-movers-
# <date>.json (fetched from stockanalysis.com). Key honesty: the raw biggest
# movers are almost always microcaps you cannot trade, so each carries a SKIP
# reason, and only names clearing the liquidity screen get the full protocol.
from market_movers import parse_num, screen, classify_untradeable  # noqa: E402

MOVER_COLS = [
    Col("side", "Gainer / Loser", 13, "Which list the stock topped today."),
    Col("sym", "Ticker", 9, "The stock symbol."),
    Col("company", "Company", 30, "Company name."),
    Col("price", "Price", 10, "Price at the close.", FMT_PRICE),
    Col("pct", "Day Change %", 12, "The move that put it on the list.", FMT_PCT),
    Col("mcap", "Market Cap", 13,
        "Company size. Standing order: only $2B+ (mid-cap and larger) appears on "
        "this tab at all - smaller names gap, halt, and cannot absorb a $100k "
        "order, so they are dropped and counted in the note below the table."),
    Col("dvol", "$ Traded Today", 14,
        "Price x volume. A $100k order needs this to be large, or you move the "
        "price against yourself getting in and out."),
    Col("verdict", "PROTOCOL VERDICT", 24,
        "What to do. SKIP = untradeable at your size (almost all of them). The few "
        "that clear the screen get a real buy/short call."),
    Col("why", "Reasoning", 78,
        "Plain-English why. For SKIPs, the specific disqualifier. For survivors, "
        "the full gate-stack read."),
]


def _load_movers():
    path = os.path.join(REPO, "watchlists", f"market-movers-{DATE}.json")
    if not os.path.exists(path):
        return None
    return json.load(open(path))


mv_data = _load_movers()
ws = wb.create_sheet("8 - Market Movers")
if not mv_data:
    style_header_row(ws, ["Market Movers"], ["No movers file for this date."],
                     [40], freeze=None, autofilter=False)
    ws.append([safe(f"No watchlists/market-movers-{DATE}.json found. Fetch the day's "
                    "gainers/losers (stockanalysis.com) and re-run.")])
else:
    # Omar 2026-07-28: only MID-CAP OR BIGGER ($2B+) belongs on this tab.
    # Anything smaller is not investable at our size, so listing forty microcap
    # pumps was noise. They are DROPPED entirely, and the drop is counted below
    # so a short list is never mistaken for a quiet day.
    MIN_INVESTABLE_CAP = 2e9
    mrows, _dropped = [], {"gainers": 0, "losers": 0}
    # liquid movers first (they got full protocol) - by definition $2B+
    for lm in mv_data.get("liquid_movers", []):
        mrows.append({
            "side": lm.get("side", "gainer").upper() + " (liquid)",
            "sym": lm["sym"], "company": lm.get("company"), "price": lm.get("price"),
            "pct": lm.get("pct_change"), "mcap": _fmt_cap(lm.get("market_cap")),
            "dvol": _fmt_cap(lm.get("avg_dollar_vol")),
            "verdict": lm.get("protocol", "NEEDS CHART"),
            "why": lm.get("protocol_why", ""),
        })
    _liquid_syms = {r["sym"] for r in mrows}
    for side_key in ("gainers", "losers"):
        for m in mv_data.get(side_key, []):
            cap = m.get("market_cap")
            if cap is None:
                cap = parse_num(m.get("market_cap_raw"))
            if m["sym"] in _liquid_syms:
                continue          # already shown with the full protocol read
            if cap is None or cap < MIN_INVESTABLE_CAP:
                _dropped[side_key] += 1
                continue
            side = m.get("side", side_key[:-1])
            if m.get("tradeable"):
                verdict, why = "TRADEABLE - see chart", \
                    "Cleared the liquidity screen; chart it and run the gate stack."
            else:
                verdict, why = classify_untradeable(m, side)
                why = f"{m.get('skip_reason','')}. {why}"
            mrows.append({
                "side": side.upper(), "sym": m["sym"], "company": m.get("company"),
                "price": m.get("price"), "pct": m.get("pct_change"),
                "mcap": _fmt_cap(cap),
                "dvol": _fmt_cap((m.get("price") or 0) * (m.get("volume") or 0)),
                "verdict": verdict, "why": why,
            })

    def _mv_fill(r):
        v = str(r.get("verdict") or "")
        if v.startswith(("BUY", "STARTER")):
            return GREEN
        if "SHORT" in v or "TRADEABLE" in v:
            return AMBER
        return GREY   # the SKIPs - visually recede so the tradeable rows pop

    write_table(ws, MOVER_COLS, mrows, autofilter=True, row_fill=_mv_fill)
    if not mrows:
        ws.append([])
        ws.append([safe("NOTHING INVESTABLE MOVED. Every one of today's biggest "
                        "gainers and losers was below $2B - microcap pumps and "
                        "crashes, none of them tradeable at our size.")])
        ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append([])
    ws.append([safe(f"Source: {mv_data.get('source','?')}, captured {mv_data.get('captured','?')}.")])
    _n_drop = _dropped["gainers"] + _dropped["losers"]
    ws.append([safe(f"STANDING ORDER (Omar 2026-07-28): only mid-cap and larger "
                    f"($2 billion+) is shown here - anything smaller is not "
                    f"investable at our size. Dropped today on that rule: "
                    f"{_dropped['gainers']} gainers and {_dropped['losers']} losers "
                    f"({_n_drop} names, mostly microcap pumps, post-halt spikes "
                    f"and penny stocks).")])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append([safe("For the big LIQUID movers that reflect real rotation, the "
                    "'1 - Market & Rotation' tab and the ranked buy list are the "
                    "actionable views - this tab is the raw market-wide scan.")])

# ------------------------------------------------------- 9 Track Record ---
# The origination scanner's own performance history (recommendations_log.csv),
# tracked since the scans began. Surfaced here per Omar 2026-07-22.
_tr = None
_trpath = os.path.join(REPO, "research", "track-record.json")
if os.path.exists(_trpath):
    _tr = json.load(open(_trpath))

# Omar 2026-07-28: this tab must read like the origination scanner's own TRACKER
# tab - every pick broken out INDIVIDUALLY, showing how it has done since the day
# it was recommended. The grade/tab averages that used to be the whole tab now
# sit below the pick list as the summary.
TRP_COLS = [
    Col("date", "Picked On", 11, "The date the origination scanner recommended it."),
    Col("ticker", "Ticker", 9, "The stock symbol."),
    Col("name", "Company", 28, "Company name."),
    Col("grade", "Grade", 8, "A+ / A / B / C from the origination scan."),
    Col("tab", "From Scan Tab", 16, "Which origination-scan tab produced the pick."),
    Col("rec_price", "Price Then", 11, "The price on the day of the call.", FMT_PRICE),
    Col("cur_price", "Price Now", 11, "The latest tracked price.", FMT_PRICE),
    Col("ret_pct", "% Since Picked", 13,
        "Return from the call to now (unrealized, the scanner's own tracking).", FMT_PCT),
    Col("best_pct", "Best % Reached", 13,
        "The most it was ever up - the upside that was on the table.", FMT_PCT),
    Col("worst_pct", "Worst % Reached", 13,
        "The most it was ever down - the pain a holder sat through.", FMT_PCT),
    Col("days_held", "Days", 7, "Days since the call.", FMT_INT),
    Col("status", "Status", 10, "OPEN = still being tracked."),
]


def _load_picks():
    """Per-pick rows: from track-record.json if present, else straight from the
    dated log snapshot in research/ (older json files predate the picks list)."""
    if _tr and _tr.get("picks"):
        return _tr["picks"]
    import csv
    import glob as _glob
    snaps = sorted(_glob.glob(os.path.join(REPO, "research",
                                           "recommendations_log-*.csv")))
    if not snaps:
        return []
    def _f(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return None
    picks = []
    for r in csv.DictReader(open(snaps[-1], encoding="utf-8", errors="replace")):
        picks.append({
            "date": r.get("date"), "ticker": r.get("ticker"), "name": r.get("name"),
            "grade": r.get("grade"), "tab": r.get("tab"),
            "rec_price": _f(r.get("price")), "cur_price": _f(r.get("cur_price")),
            "ret_pct": round(_f(r.get("cur_ret")) * 100, 2)
                       if _f(r.get("cur_ret")) is not None else None,
            "best_pct": round(_f(r.get("best_ret")) * 100, 1)
                        if _f(r.get("best_ret")) is not None else None,
            "worst_pct": round(_f(r.get("worst_ret")) * 100, 1)
                         if _f(r.get("worst_ret")) is not None else None,
            "days_held": _f(r.get("days_held")), "status": r.get("status"),
        })
    picks.sort(key=lambda p: (p["date"] or "", p["ticker"] or ""), reverse=True)
    return picks


# ---- MY ACTIONED TRADES (Omar 2026-07-28): the recommendations he actually
# put money on, tracked as their own tally beside the scanner's pick history.
# Fed by research/calls-ledger.json (fills + close blocks - the standing rule
# is that every reported fill/close updates the ledger and open-book.json, so
# this table maintains itself). First entries: DVN (closed 7/28) + HAS/SJM/MLI.
ACTIONED_COLS = [
    Col("sym", "MY TRADES - Ticker", 16,
        "A recommendation Omar actually traded, from the calls ledger. This "
        "tally measures the ACTED-ON book, separate from the scanner's "
        "hypothetical pick history to the left."),
    Col("opened", "Opened", 11, "The date the position was put on."),
    Col("status", "Status", 14, "OPEN, or CLOSED with the exit date."),
    Col("shares", "Shares", 8, "Position size in shares.", FMT_INT),
    Col("entry", "Entry", 10, "Actual fill price.", FMT_PRICE),
    Col("stop", "Stop", 10, "The live stop (open positions) or '-' once closed.", FMT_PRICE),
    Col("now", "Now / Exit", 11,
        "Latest scan price for open positions; the actual exit fill for closed "
        "ones.", FMT_PRICE),
    Col("pnl", "P&L $", 11,
        "Dollars made or lost. Open positions = unrealized (paper); closed = "
        "banked.", FMT_INT),
    Col("pnl_pct", "P&L %", 9, "The same, as a percent of entry.", FMT_PCT),
]


def _actioned_rows():
    try:
        led = json.load(open(os.path.join(REPO, "research", "calls-ledger.json"),
                             encoding="utf-8"))
        calls = led.get("calls", [])
    except Exception:
        return []
    try:
        book = {p["sym"]: p for p in json.load(
            open(os.path.join(REPO, "watchlists", "open-book.json"),
                 encoding="utf-8")).get("positions", [])}
    except Exception:
        book = {}
    out = []
    for c in calls:
        m = re.search(r"FILLED\s+([\d,]+)\s*sh\s*@\s*\$?([\d.]+)",
                      str(c.get("fill") or ""))
        if not m:
            continue          # analysis-only call, never traded
        shares, entry = int(m.group(1).replace(",", "")), float(m.group(2))
        sym, cl = c.get("sym"), c.get("close") or {}
        if cl:                # closed - the ledger close block is the truth
            now, status, stop = cl.get("fill"), f"CLOSED {cl.get('date', '')}", "-"
            pnl = cl.get("pnl")
            if pnl is None and now:
                pnl = round((now - entry) * shares)
        else:                 # open - live stop from the book, price from today's scan
            b = book.get(sym, {})
            a = allrows.get(sym, {})
            stop = b.get("stop") or c.get("stop")
            # Scan real quote first; public-scanner close as fallback (2026-07-28:
            # SJM's sweep quote came back empty and its P&L cell shipped blank).
            now = (a.get("real_last") or a.get("real_close")
                   or TV_METRICS.get(sym, {}).get("close"))
            status = "OPEN"
            pnl = round((now - entry) * shares) if now else None
        out.append({"sym": sym, "opened": c.get("date"), "status": status,
                    "shares": shares, "entry": entry, "stop": stop, "now": now,
                    "pnl": pnl,
                    "pnl_pct": round((now / entry - 1) * 100, 1) if now else None})
    out.sort(key=lambda r: (r["status"] != "OPEN", r["opened"] or ""))
    return out


ws = wb.create_sheet("9 - Track Record")
_picks = _load_picks()
if (not _tr or _tr.get("error")) and not _picks:
    style_header_row(ws, ["Track Record"], ["Run track_record.py first."],
                     [40], freeze=None, autofilter=False)
    ws.append([safe("No research/track-record.json - run scripts/track_record.py.")])
else:
    def _pick_fill(r):
        m = r.get("ret_pct")
        if m is None:
            return None
        return GREEN if m > 0 else (RED if m < -3 else AMBER)

    write_table(ws, TRP_COLS, _picks, autofilter=True, row_fill=_pick_fill)

    _acted = _actioned_rows()
    if _acted:
        _ac_start = len(TRP_COLS) + 2          # one blank gutter column
        _last = side_table(
            ws, ACTIONED_COLS, _acted, _ac_start,
            row_fill=lambda r: (GREEN if (r.get("pnl") or 0) > 0
                                else RED if (r.get("pnl") or 0) < 0 else None))
        _real = sum(r["pnl"] for r in _acted
                    if r["pnl"] is not None and not r["status"].startswith("OPEN"))
        _unreal = sum(r["pnl"] for r in _acted
                      if r["pnl"] is not None and r["status"].startswith("OPEN"))
        _sum_cell = ws.cell(row=_last + 2, column=_ac_start, value=safe(
            f"TALLY: banked {_real:+,.0f} on closed trades; open positions "
            f"marking {_unreal:+,.0f} (paper, not banked)."))
        _sum_cell.font = BOLD
        ws.cell(row=_last + 3, column=_ac_start, value=safe(
            "This block tracks what Omar actually TRADED from the "
            "recommendations - the left table is every pick, acted on or not. "
            "Updates itself from the calls ledger and the open book."))

    ws.append([])
    ws.append([safe("SCORECARD - the averages behind the list above. Green share = "
                    "how many picks are currently in profit.")])
    ws.cell(row=ws.max_row, column=1).font = SECTION
    if _tr and not _tr.get("error"):
        o = _tr["overall"]
        ws.append([safe(f"ALL {o['n']} PICKS: {o['win_pct']}% green, average "
                        f"{o['mean']:+.1f}%, median {o['median']:+.1f}%, average "
                        f"best point +{o['mfe']}%, average worst point {o['mae']}%.")])
        ws.cell(row=ws.max_row, column=1).font = BOLD
        for g, v in _tr["by_grade"].items():
            ws.append([safe(f"  grade {g}: {v['n']} picks, {v['win_pct']}% green, "
                            f"average {v['mean']:+.1f}%, median {v['median']:+.1f}%")])
        for t, v in _tr["by_tab"].items():
            ws.append([safe(f"  tab {t}: {v['n']} picks, {v['win_pct']}% green, "
                            f"average {v['mean']:+.1f}%, median {v['median']:+.1f}%")])
    if _tr and not _tr.get("error"):
        ws.append([])
        ws.append([safe(f"ORIGINATION SCANNER TRACK RECORD - {_tr['total']} picks, "
                        f"{_tr['date_range'][0]} to {_tr['date_range'][1]}. Returns are "
                        f"UNREALIZED (matured {o['matured']}/{_tr['total']} at the 21-day "
                        f"mark; the log only started 7/1). This grades the scanner's own "
                        f"A/B/C picks, separate from the trade-call ledger.")])
        ws.cell(row=ws.max_row, column=1).font = BOLD
        q = _tr.get("score_quartile")
        if q:
            ws.append([safe(f"DOES THE SCORE PREDICT? bottom-25% by score mean "
                            f"{q['bottom_q_mean']:+.2f}%, top-25% mean {q['top_q_mean']:+.2f}%, "
                            f"spread {q['spread']:+.2f}pt.")])
            ws.cell(row=ws.max_row, column=1).font = RED_FONT if q["spread"] <= 0 else BOLD
            if q["spread"] <= 0:
                ws.append([safe("The origination score does NOT predict returns here - flat/"
                                "backwards, the SAME result as the BUY SCORE calibration. Two "
                                "independent scores, same verdict: sort what to look at, do "
                                "not forecast.")])
        ws.append([safe("These are unrealized reads on OPEN positions, not final "
                        "results. Check the grade lines above before assuming a "
                        "higher grade is worth more - so far it has not been.")])

# ---------------------------------------------------------- 10 Data Quality ---
def _dq_polish(d):
    """Every data-quality row reads plainly and no column is left blank.

    Omar 2026-07-28: 'Effect on this run' and 'What to do' arrived empty on
    INFO rows and the findings carried indicator shorthand. An empty impact
    cell reads as 'unknown damage', which is worse than saying 'none'.
    """
    d = dict(d)
    sev = str(d.get("severity", "INFO")).upper()
    d["finding"] = plainify(d.get("finding") or "")
    if not d.get("impact"):
        d["impact"] = {
            "BLOCKER": "Until this is fixed, conclusions in this workbook may "
                       "simply be wrong - treat the affected numbers as suspect.",
            "WARNING": "The numbers stand, but do not put money on the affected "
                       "names before running the check in the next column.",
        }.get(sev, "No damage to the numbers - this is context you should have "
                   "while reading, not a defect.")
    else:
        d["impact"] = plainify(d["impact"])
    if not d.get("action"):
        d["action"] = ("Fix it before trusting this run." if sev == "BLOCKER"
                       else "Confirm on the live chart before acting."
                       if sev == "WARNING" else "Nothing to do - context only.")
    else:
        d["action"] = plainify(d["action"])
    return d


dq = [_dq_polish(classify(q) if isinstance(q, str) else q)
      for q in res.get("data_quality", [])]
if _TV_METRICS_NOTE:
    dq.append(_dq_polish({
        "severity": "WARNING", "area": "Data feed",
        "finding": ("The public-scanner fetch for RSI, volume-vs-normal and "
                    f"52-week-high distance failed ({_TV_METRICS_NOTE}), so those "
                    "three Fresh Buys columns are blank this run."),
        "impact": "", "action": "Re-run the compile with network access to fill them."}))
data_quality_sheet(wb.create_sheet("10 - Data Quality"), dq)

# ------------------------------------------------------------ 9 Run Summary ---
SUM_COLS = [
    Col("field", "Field", 30, "The run attribute."),
    Col("value", "Value", 110, "Its value for this run."),
]
sum_rows = [
    {"field": "Scan data as of", "value": f"{DATE} close (live chart reads)"},
    {"field": "Scanner enrich as of", "value": ENRICH_AS_OF},
    {"field": "Workbook written", "value": f"{WRITTEN} local"},
    {"field": "Date / variant / system", "value":
        f"{res.get('date')} / {res.get('variant')} / {res.get('system_version')}"},
    {"field": "Universe / scanned / fresh", "value":
        f"{res.get('universe_size')} / {res.get('scanned')} / {res.get('fresh_count')}"},
    {"field": "Sell-mode count", "value": len(res.get("sell_mode", []))},
    {"field": "Blocked count", "value": len(block_rows)},
    {"field": "Combined risk", "value": res.get("combined_risk_new_actives", "")},
    {"field": "Data-quality BLOCKERS", "value":
        sum(1 for d in dq if str(d.get("severity")).upper() == "BLOCKER")},
    {"field": "Run note", "value": res.get("run_note", "")},
]
ws = wb.create_sheet("11 - Run Summary")
write_table(ws, SUM_COLS, sum_rows, autofilter=False)
for n in range(2, ws.max_row + 1):
    ws.row_dimensions[n].height = 44

# Save, stepping around whatever is currently open in Excel. The first fallback
# was a single "_new" name, which broke the moment BOTH files were open - the
# fallback itself raised. Keep trying until something lands.
_stamp = datetime.datetime.now().strftime("%H%M%S")
_candidates = [f"universe_{DATE}.xlsx", f"universe_{DATE}_new.xlsx",
               f"universe_{DATE}_{_stamp}.xlsx"]
if OUT_OVERRIDE:
    _candidates = [OUT_OVERRIDE]
out, _err = None, None
for _name in _candidates:
    _path = _name if os.path.isabs(_name) else os.path.join(REPO, "reports", _name)
    try:
        wb.save(_path)
        out = _path
        break
    except PermissionError as exc:
        _err = exc
        continue
if out is None:
    raise SystemExit(f"could not write any workbook - all candidates locked: {_err}")
_first = (_candidates[0] if os.path.isabs(_candidates[0])
          else os.path.join(REPO, "reports", _candidates[0]))
if out != _first:
    print(f"NOTE: earlier filename(s) locked in Excel - wrote {os.path.basename(out)} instead.")
print(f"written {out}")
print(f"  sheets: {len(wb.sheetnames)} | fresh {len(hits)} | blocked {len(block_rows)} "
      f"| notes {len(note_rows)} | dq {len(dq)} "
      f"(BLOCKERS {sum(1 for d in dq if str(d.get('severity')).upper() == 'BLOCKER')})")
